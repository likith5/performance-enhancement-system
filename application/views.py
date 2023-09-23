from application import create_app
from flask import Blueprint,url_for,session,flash,make_response
from flask_session import Session
from application import db
from pymongo.errors import PyMongoError
from flask import render_template,request,redirect,Response
from bson.binary import Binary
from io import StringIO
import openpyxl
from openpyxl.styles import PatternFill
import json



# this is for xls writer 
import xlsxwriter

views = Blueprint('views',__name__)

# routing has been divided into  two sub categories one is views.py and the other is auth.py views.py contains all the functionalites regarding the templates
#  on how the renderingand routing of each page takes place which will be mentioned in the each comments below.....
@views.route("/")
def index():
    # index acts as a base layout page where nav bar and other design templates or rendered commonly 
    if 'email' in session:
        return redirect(url_for('views.dashboard'))
    return redirect(url_for('auth.login'))
    

    
    return render_template('index.html')

# profile page containes the deatils of all the individiual student
@views.route("/profile")
def profile():
    
# it has been divided into two sub categories one is for teacher and the other is for student which willl be in session
    title="profile"
    if  'teacheremail' in session:
        # session will be obtained and used to query out the data from the database and used to render it on the front end
        usn = session.get('studentname')
        user = db.users.find_one({'usn':usn })
        username = user["personal"].get('username')
        email = user["personal"].get('email')
        college = user["personal"].get('college')
        sem = user["personal"].get('sem')
        linkdin = user["personal"].get('linkdin')
        github = user["personal"].get('github')
        project1 = user["personal"].get('project-1')
        intership1 = user["personal"].get('internship-1')
        interst1 = user["personal"].get('interst-1')
        title="dashboard"
        return render_template('profile.html',title=title,usn=usn,username=username,email=email,college=college,sem=sem,linkdin=linkdin,github=github,intership1=intership1,project1=project1,interst1=interst1)
   
    if 'studentemail' in session :
        usn = session.get('studentemail')
        user = db.users.find_one({'usn':usn })
        username = user["personal"].get('username')
        userusn=user["usn"]
        email = user["personal"].get('email')
        college = user["personal"].get('college')
        phone = user["personal"].get('phone')
        linkdin = user["personal"].get('linkdin')
        github = user["personal"].get('github')
        
        title="dashboard"

        


        Assessment1 =user.get('Assessment1')
        Assessment2 =user.get('Assessment2')
        Assessment3 =user.get('Assessment3')
        Assessment4 =user.get('Assessment4')
        Assessment5 =user.get('Assessment5')
        Assessment6 =user.get('Assessment6')
        Assessment7 =user.get('Assessment7')
        Assessment8 =user.get('Assessment8')
    #    for each assessment there has been using individual profiling system which is based on the marks enterd by teachers so for each assessments availible there 
    # will be seperate profiling based on the test
    # this is for test1.... 
        
        if   Assessment1 is not None and Assessment2 is  None  and Assessment3 is  None  and Assessment4 is  None  and Assessment5  is  None and Assessment6  is  None and Assessment7 is  None and Assessment8  is  None:
        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            
            def summary(userr):
                p=[]
                for key, value in userr.items():
                        p.append(int(value))
                return p  


            list_of_first_test_results=summary(test1_skills)


            

            communication=[list_of_first_test_results[0]]
            technical=[list_of_first_test_results[1]]
            creativity=[list_of_first_test_results[2]]
            projectmmt=[list_of_first_test_results[3]]
            timemmt=[list_of_first_test_results[4]]            
            gk=[list_of_first_test_results[5]]     
            interpersonal=[list_of_first_test_results[6]]
            resultoriented=[list_of_first_test_results[7]]
            leadership=[list_of_first_test_results[8]]
            presentation=[list_of_first_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10]]
            aptitude=[list_of_first_test_results[11]]
            # communication1 = int(user['Assessment1'].get('communication'))
            # technical1 = int(user["Assessment1"].get('technical'))
            # creativity1 = int(user['Assessment1'].get('creativity'))
            # projectmm1 = int(user['Assessment1'].get('projectmmt'))
            # timemanagement1 = int(user['Assessment1'].get('timemanagement'))
            # generalknowledge1 = int(user['Assessment1'].get('generalknowledge'))
            # interpersonal1 = int(user['Assessment1'].get('interpersonal'))
            # resultoriented1 = int(user['Assessment1'].get('resultoriented'))
            # leardership1 = int(user['Assessment1'].get('leardership'))
            # presentation1 = int(user['Assessment1'].get('presentation'))

            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/15)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/25)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/25)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/25)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/25)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/25)*5))
            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)
            
            print(aptitude)
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('profile.html',title=title,usn=usn,user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,username=username,email=email,college=college,phone=phone,linkdin=linkdin,github=github,userusn=userusn)
        # this is for test2..  
        elif  Assessment1 is not None and Assessment2 is not None and  Assessment3 is  None and Assessment4 is  None and Assessment5 is  None and Assessment6 is  None  and Assessment7 is  None and Assessment8 is  None:
            
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            def summary(userr):
                p=[]
                for key, value in userr.items():
                    p.append(int(value))
                return p  
                        
                                

                        

          
            
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
           
         
            

            communication=[list_of_first_test_results[0],list_of_second_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4]]            
            gk=[list_of_first_test_results[5],list_of_second_test_results[5]]     
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11]]

            # communication1 = int(user['Assessment1'].get('communication'))
            # technical1 = int(user["Assessment1"].get('technical'))
            # creativity1 = int(user['Assessment1'].get('creativity'))
            # projectmm1 = int(user['Assessment1'].get('projectmmt'))
            # timemanagement1 = int(user['Assessment1'].get('timemanagement'))
            # generalknowledge1 = int(user['Assessment1'].get('generalknowledge'))
            # interpersonal1 = int(user['Assessment1'].get('interpersonal'))
            # resultoriented1 = int(user['Assessment1'].get('resultoriented'))
            # leardership1 = int(user['Assessment1'].get('leardership'))
            # presentation1 = int(user['Assessment1'].get('presentation'))

            

            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/30)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/50)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/50)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/50)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/50)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/50)*5))

            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)
           
            print(aptitude)
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('profile.html',title=title,usn=usn,user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,username=username,email=email,college=college,phone=phone,linkdin=linkdin,github=github,userusn=userusn)
    #    this is for test3 ... 
        elif    Assessment1 is not None and Assessment2 is not None and  Assessment3 is not None and Assessment4 is  None and Assessment5 is  None and Assessment6 is  None  and Assessment7 is  None and Assessment8 is  None:


        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            
            def summary(userr):
                p=[]
                for key, value in userr.items():
                        p.append(int(value))
                return p  
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            test3_skills = user["Assessment3"]
          
            
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
            list_of_third_test_results=summary(test3_skills)
           
         
            

            communication=[list_of_first_test_results[0],list_of_second_test_results[0],list_of_third_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1],list_of_third_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2],list_of_third_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3],list_of_third_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4],list_of_third_test_results[4]]            
            gk=[list_of_first_test_results[5],list_of_second_test_results[5],list_of_third_test_results[5]]     
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6],list_of_third_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7],list_of_third_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8],list_of_third_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9],list_of_third_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10],list_of_third_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11],list_of_third_test_results[11]]
            print(aptitude)

            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/45)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/75)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/75)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/75)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/60)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/75)*5))

            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)
            
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('profile.html',title=title,usn=usn,user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,username=username,email=email,college=college,phone=phone,linkdin=linkdin,github=github,userusn=userusn)
    #    this is for assessmet4 ....

        elif    Assessment1 is not None and Assessment2 is not None and  Assessment3 is not None and Assessment4 is not  None and Assessment5 is  None and Assessment6 is  None  and Assessment7 is  None and Assessment8 is  None:


        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            test3_skills = user["Assessment3"]
            test4_skills = user["Assessment4"]
            def summary(userr):
                        p=[]
                        for key, value in userr.items():
                                p.append(int(value))
                        return p  
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
            list_of_third_test_results=summary(test3_skills)
            list_of_fourth_test_results=summary(test4_skills)
          
          

            communication=[list_of_first_test_results[0],list_of_second_test_results[0],list_of_third_test_results[0],list_of_fourth_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1],list_of_third_test_results[1],list_of_fourth_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2],list_of_third_test_results[2],list_of_fourth_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3],list_of_third_test_results[3],list_of_fourth_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4],list_of_third_test_results[4],list_of_fourth_test_results[4]]
            gk=[list_of_first_test_results[5],list_of_second_test_results[5],list_of_third_test_results[5],list_of_fourth_test_results[5]]
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6],list_of_third_test_results[6],list_of_fourth_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7],list_of_third_test_results[7],list_of_fourth_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8],list_of_third_test_results[8],list_of_fourth_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9],list_of_third_test_results[9],list_of_fourth_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10],list_of_third_test_results[10],list_of_fourth_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11],list_of_third_test_results[11],list_of_fourth_test_results[11]]
            print(aptitude)
            
            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/60)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/100)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/100)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/100)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/80)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/100)*5))
            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)

            
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('profile.html',title=title,usn=usn,user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,username=username,email=email,college=college,phone=phone,linkdin=linkdin,github=github,userusn=userusn)
        # this is for test5...
        elif  Assessment1 is not None and Assessment2 is not None and  Assessment3 is not None and Assessment4 is not  None and Assessment5 is not  None and Assessment6 is  None  and Assessment7 is  None and Assessment8 is  None:


        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            test3_skills = user["Assessment3"]
            test4_skills = user["Assessment4"]
            test5_skills = user["Assessment5"]
            def summary(userr):
                p=[]
                for key, value in userr.items():
                        p.append(int(value))
                return p  
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
            list_of_third_test_results=summary(test3_skills)
            list_of_fourth_test_results=summary(test4_skills)
            list_of_fifth_test_results=summary(test5_skills)
          
          

            communication=[list_of_first_test_results[0],list_of_second_test_results[0],list_of_third_test_results[0],list_of_fourth_test_results[0],list_of_fifth_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1],list_of_third_test_results[1],list_of_fourth_test_results[1],list_of_fifth_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2],list_of_third_test_results[2],list_of_fourth_test_results[2],list_of_fifth_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3],list_of_third_test_results[3],list_of_fourth_test_results[3],list_of_fifth_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4],list_of_third_test_results[4],list_of_fourth_test_results[4],list_of_fifth_test_results[4]]
            gk=[list_of_first_test_results[5],list_of_second_test_results[5],list_of_third_test_results[5],list_of_fourth_test_results[5],list_of_fifth_test_results[5]]
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6],list_of_third_test_results[6],list_of_fourth_test_results[6],list_of_fifth_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7],list_of_third_test_results[7],list_of_fourth_test_results[7],list_of_fifth_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8],list_of_third_test_results[8],list_of_fourth_test_results[8],list_of_fifth_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9],list_of_third_test_results[9],list_of_fourth_test_results[9],list_of_fifth_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10],list_of_third_test_results[10],list_of_fourth_test_results[10],list_of_fifth_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11],list_of_third_test_results[11],list_of_fourth_test_results[11],list_of_fifth_test_results[11]]
            print(aptitude)
            
            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/75)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/125)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/125)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/125)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/100)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/125)*5))
    
            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)

            
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('profile.html',title=title,usn=usn,user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,username=username,email=email,college=college,phone=phone,linkdin=linkdin,github=github,userusn=userusn)
    #    this is for test6... 
        elif  Assessment1 is not None and Assessment2 is not None and  Assessment3 is not None and Assessment4 is not  None and Assessment5 is not  None and Assessment6 is  not None  and Assessment7 is  None and Assessment8 is  None:


        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            test3_skills = user["Assessment3"]
            test4_skills = user["Assessment4"]
            test5_skills = user["Assessment5"]
            test6_skills = user["Assessment6"]
            def summary(userr):
                p=[]
                for key, value in userr.items():
                        p.append(int(value))
                return p  
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
            list_of_third_test_results=summary(test3_skills)
            list_of_fourth_test_results=summary(test4_skills)
            list_of_fifth_test_results=summary(test5_skills)
            list_of_sixth_test_results=summary(test6_skills)
            # communication1 = int(user['Assessment1'].get('communication'))
            # technical1 = int(user["Assessment1"].get('technical'))
            # creativity1 = int(user['Assessment1'].get('creativity'))
            # projectmm1 = int(user['Assessment1'].get('projectmmt'))
            # timemanagement1 = int(user['Assessment1'].get('timemanagement'))
            # generalknowledge1 = int(user['Assessment1'].get('generalknowledge'))
            # interpersonal1 = int(user['Assessment1'].get('interpersonal'))
            # resultoriented1 = int(user['Assessment1'].get('resultoriented'))
            # leardership1 = int(user['Assessment1'].get('leardership'))
            # presentation1 = int(user['Assessment1'].get('presentation'))
            communication=[list_of_first_test_results[0],list_of_second_test_results[0],list_of_third_test_results[0],list_of_fourth_test_results[0],list_of_fifth_test_results[0],list_of_sixth_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1],list_of_third_test_results[1],list_of_fourth_test_results[1],list_of_fifth_test_results[1],list_of_sixth_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2],list_of_third_test_results[2],list_of_fourth_test_results[2],list_of_fifth_test_results[2],list_of_sixth_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3],list_of_third_test_results[3],list_of_fourth_test_results[3],list_of_fifth_test_results[3],list_of_sixth_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4],list_of_third_test_results[4],list_of_fourth_test_results[4],list_of_fifth_test_results[4],list_of_sixth_test_results[4]]
            gk=[list_of_first_test_results[5],list_of_second_test_results[5],list_of_third_test_results[5],list_of_fourth_test_results[5],list_of_fifth_test_results[5],list_of_sixth_test_results[5]]
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6],list_of_third_test_results[6],list_of_fourth_test_results[6],list_of_fifth_test_results[6],list_of_sixth_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7],list_of_third_test_results[7],list_of_fourth_test_results[7],list_of_fifth_test_results[7],list_of_sixth_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8],list_of_third_test_results[8],list_of_fourth_test_results[8],list_of_fifth_test_results[8],list_of_sixth_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9],list_of_third_test_results[9],list_of_fourth_test_results[9],list_of_fifth_test_results[9],list_of_sixth_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10],list_of_third_test_results[10],list_of_fourth_test_results[10],list_of_fifth_test_results[10],list_of_sixth_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11],list_of_third_test_results[11],list_of_fourth_test_results[11],list_of_fifth_test_results[11],list_of_sixth_test_results[11]]
            print(sum(technical))
            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/90)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/150)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/150)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/150)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/120)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/150)*5))
            print(customeri)
            print(projecti)
            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)

            
            

            
            print('++++')
            return render_template('profile.html',title=title,usn=usn,user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,username=username,email=email,college=college,phone=phone,linkdin=linkdin,github=github,userusn=userusn)
        # this is for test7
        elif  Assessment1 is not None and Assessment2 is not None and  Assessment3 is not None and Assessment4 is not  None and Assessment5 is not None and Assessment6 is not None  and Assessment7 is not None and Assessment8 is  None:


        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            test3_skills = user["Assessment3"]
            test4_skills = user["Assessment4"]
            test5_skills = user["Assessment5"]
            test6_skills = user["Assessment6"]
            test7_skills = user["Assessment7"]
            def summary(userr):
                p=[]
                for key, value in userr.items():
                        p.append(int(value))
                return p  
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
            list_of_third_test_results=summary(test3_skills)
            list_of_fourth_test_results=summary(test4_skills)
            list_of_fifth_test_results=summary(test5_skills)
            list_of_sixth_test_results=summary(test6_skills)
            list_of_seventh_test_results=summary(test7_skills)
           
           

            communication=[list_of_first_test_results[0],list_of_second_test_results[0],list_of_third_test_results[0],list_of_fourth_test_results[0],list_of_fifth_test_results[0],list_of_sixth_test_results[0],list_of_seventh_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1],list_of_third_test_results[1],list_of_fourth_test_results[1],list_of_fifth_test_results[1],list_of_sixth_test_results[1],list_of_seventh_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2],list_of_third_test_results[2],list_of_fourth_test_results[2],list_of_fifth_test_results[2],list_of_sixth_test_results[2],list_of_seventh_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3],list_of_third_test_results[3],list_of_fourth_test_results[3],list_of_fifth_test_results[3],list_of_sixth_test_results[3],list_of_seventh_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4],list_of_third_test_results[4],list_of_fourth_test_results[4],list_of_fifth_test_results[4],list_of_sixth_test_results[4],list_of_seventh_test_results[4]]
            gk=[list_of_first_test_results[5],list_of_second_test_results[5],list_of_third_test_results[5],list_of_fourth_test_results[5],list_of_fifth_test_results[5],list_of_sixth_test_results[5],list_of_seventh_test_results[5]]
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6],list_of_third_test_results[6],list_of_fourth_test_results[6],list_of_fifth_test_results[6],list_of_sixth_test_results[6],list_of_seventh_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7],list_of_third_test_results[7],list_of_fourth_test_results[7],list_of_fifth_test_results[7],list_of_sixth_test_results[7],list_of_seventh_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8],list_of_third_test_results[8],list_of_fourth_test_results[8],list_of_fifth_test_results[8],list_of_sixth_test_results[8],list_of_seventh_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9],list_of_third_test_results[9],list_of_fourth_test_results[9],list_of_fifth_test_results[9],list_of_sixth_test_results[9],list_of_seventh_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10],list_of_third_test_results[10],list_of_fourth_test_results[10],list_of_fifth_test_results[10],list_of_sixth_test_results[10],list_of_seventh_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11],list_of_third_test_results[11],list_of_fourth_test_results[11],list_of_fifth_test_results[11],list_of_sixth_test_results[11],list_of_seventh_test_results[11]]
            print(aptitude)
            
            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/105)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/175)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/175)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/175)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/140)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/175)*5))
            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)

            
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('profile.html',title=title,usn=usn,user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,username=username,email=email,college=college,phone=phone,linkdin=linkdin,github=github,userusn=userusn)
        # this is for test8
        elif Assessment8 is not None:
        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            test3_skills = user["Assessment3"]
            test4_skills = user["Assessment4"]
            test5_skills = user["Assessment5"]
            test6_skills = user["Assessment6"]
            test7_skills = user["Assessment7"]
            test8_skills = user["Assessment8"]
            def summary(userr):
                p=[]
                for key, value in userr.items():
                        p.append(int(value))
                return p  
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
            list_of_third_test_results=summary(test3_skills)
            list_of_fourth_test_results=summary(test4_skills)
            list_of_fifth_test_results=summary(test5_skills)
            list_of_sixth_test_results=summary(test6_skills)
            list_of_seventh_test_results=summary(test7_skills)
            print(list_of_sixth_test_results)
            list_of_eight_test_results=summary(test8_skills)
            print(list_of_eight_test_results)
            list_of_assessments=["Assessment1","Assessment2","Assessment3","Assessment4","Assessment5","Assessment6","Assessment7","Assessment8"]
            print(list_of_assessments)

            communication=[list_of_first_test_results[0],list_of_second_test_results[0],list_of_third_test_results[0],list_of_fourth_test_results[0],list_of_fifth_test_results[0],list_of_sixth_test_results[0],list_of_seventh_test_results[0],list_of_eight_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1],list_of_third_test_results[1],list_of_fourth_test_results[1],list_of_fifth_test_results[1],list_of_sixth_test_results[1],list_of_seventh_test_results[1],list_of_eight_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2],list_of_third_test_results[2],list_of_fourth_test_results[2],list_of_fifth_test_results[2],list_of_sixth_test_results[2],list_of_seventh_test_results[2],list_of_eight_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3],list_of_third_test_results[3],list_of_fourth_test_results[3],list_of_fifth_test_results[3],list_of_sixth_test_results[3],list_of_seventh_test_results[3],list_of_eight_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4],list_of_third_test_results[4],list_of_fourth_test_results[4],list_of_fifth_test_results[4],list_of_sixth_test_results[4],list_of_seventh_test_results[4],list_of_eight_test_results[4]]
            gk=[list_of_first_test_results[5],list_of_second_test_results[5],list_of_third_test_results[5],list_of_fourth_test_results[5],list_of_fifth_test_results[5],list_of_sixth_test_results[5],list_of_seventh_test_results[5],list_of_eight_test_results[5]]
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6],list_of_third_test_results[6],list_of_fourth_test_results[6],list_of_fifth_test_results[6],list_of_sixth_test_results[6],list_of_seventh_test_results[6],list_of_eight_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7],list_of_third_test_results[7],list_of_fourth_test_results[7],list_of_fifth_test_results[7],list_of_sixth_test_results[7],list_of_seventh_test_results[7],list_of_eight_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8],list_of_third_test_results[8],list_of_fourth_test_results[8],list_of_fifth_test_results[8],list_of_sixth_test_results[8],list_of_seventh_test_results[8],list_of_eight_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9],list_of_third_test_results[9],list_of_fourth_test_results[9],list_of_fifth_test_results[9],list_of_sixth_test_results[9],list_of_seventh_test_results[9],list_of_eight_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10],list_of_third_test_results[10],list_of_fourth_test_results[10],list_of_fifth_test_results[10],list_of_sixth_test_results[10],list_of_seventh_test_results[10],list_of_eight_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11],list_of_third_test_results[11],list_of_fourth_test_results[11],list_of_fifth_test_results[11],list_of_sixth_test_results[11],list_of_seventh_test_results[11],list_of_eight_test_results[11]]
            print(aptitude)
            
            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/120)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/200)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/200)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/200)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/160)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/200)*5))
            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)

            
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('profile.html',title=title,usn=usn,user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,username=username,email=email,college=college,phone=phone,linkdin=linkdin,github=github,userusn=userusn)

        
        

 


    flash('Please login in to view profile', category='error')

    return redirect(url_for('auth.login'))




@views.route("/dashboard",methods=['GET','POST'])
def dashboard():
    if 'studentemail' in session :
       
        usn = session.get("studentemail")
        user = db.users.find_one({'usn':usn })
        username = user["personal"].get('username')
        test11 ="Test1"
        title="dashboard"
        test1_skills = user["Assessment1"]
        
        # for item in test:
        #     print(item.)
        def summary(userr):
            p=[]
            for key, value in userr.items():
                p.append(int(value))
            return p
        test_function_summary=summary(test1_skills) 
       
        # print((max(test_function_summary)).index())
        
       
        particular_testsummary=int((sum(test_function_summary)/60)*5)
        skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
        elements={key:value for key,value in zip(skilllist,test_function_summary)}
        # Sort the dictionary based on values (descending order)
        sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
        iterator = iter(sortedinorder.items())
        first_element=next(iterator)
        # next(iterator)  # Skip the first element
        second_element = next(iterator)
        # next(iterator) 
        third_element = next(iterator)

       
        # print(sorted_dict_descending)

        # print(elements)
        
        
     
        
        if request.method=="POST":
            man= request.form.get('test')
            test1 = man.capitalize()
            test =man
            title="dashboard"
            usn = session.get("studentemail")
            user = db.users.find_one({'usn':usn })
            username = user["personal"].get('username')

            show_graph1=False
            show_graph2=False
            show_graph3=False
            show_graph4=False
            show_graph5=False
            show_graph6=False
            show_graph7=False
            show_graph8=False
            

            

            # print(test)

            if test=="Assessment1":
                if test is not None:
                
                

                
             
                    communication = int(user['Assessment1'].get('communication'))
                    technical = int(user["Assessment1"].get('technical'))
                    creativity = int(user['Assessment1'].get('creativity'))
                    projectmm = int(user['Assessment1'].get('projectmmt'))
                    timemanagement = int(user['Assessment1'].get('timemanagement'))
                    generalknowledge = int(user['Assessment1'].get('generalknowledge'))
                    interpersonal = int(user['Assessment1'].get('interpersonal'))
                    resultoriented = int(user['Assessment1'].get('resultoriented'))
                    leardership = int(user['Assessment1'].get('leardership'))
                    presentation = int(user['Assessment1'].get('presentation'))
                    entrepreneur = int(user['Assessment1'].get('entrepreneur'))
                    aptitude = int(user['Assessment1'].get('aptitude'))

                    show_graph1=True
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p
                    p=summary(test1_skills) 
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist,p)}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)

                    print(p)
                    
                    
                    particular_test1summary=int((sum(p)/60)*5)
                    return render_template('dashboard.html',show_graph1=show_graph1,first_element=first_element,second_element=second_element,third_element=third_element,particular_testsummary=particular_test1summary,test1=test,entrepreneur=entrepreneur,aptitude=aptitude,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation)
                else:
                    flash('Marks has not been allotted.', category='error')                # redirect(url_for('views.dashboard'))
            
            if test=="Assessment2":
                test2 =user.get('Assessment2')
                if test is not None:
                    test2 =user.get('Assessment2')

                
                
                    communication = int(user['Assessment2'].get('communication'))
                    technical = int(user["Assessment2"].get('technical'))
                    creativity = int(user['Assessment2'].get('creativity'))
                    projectmm = int(user['Assessment2'].get('projectmmt'))
                    timemanagement = int(user['Assessment2'].get('timemanagement'))
                    generalknowledge = int(user['Assessment2'].get('generalknowledge'))
                    interpersonal = int(user['Assessment2'].get('interpersonal'))
                    resultoriented = int(user['Assessment2'].get('resultoriented'))
                    leardership = int(user['Assessment2'].get('leardership'))
                    presentation = int(user['Assessment2'].get('presentation'))
                    entrepreneur = int(user['Assessment2'].get('entrepreneur'))
                    aptitude = int(user['Assessment2'].get('aptitude'))
                    show_graph2=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p                
                    list_of_second_test_results=summary(test2_skills)            
                    list_of_first_test_results=summary(test1_skills)            
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist,list_of_second_test_results)}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)
                
                    return render_template('dashboard.html',show_graph2=show_graph2,first_element=first_element,second_element=second_element,third_element=third_element,particular_testsummary=particular_test2summary,p=json.dumps(list_of_first_test_results),q=json.dumps(list_of_second_test_results),entrepreneur=entrepreneur,aptitude=aptitude,test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation)
                else:
                    flash('Marks has not been allotted.', category='error')                # redirect(url_for('views.dashboard'))
  
            if test=="Assessment3":
                test3 =user.get('Assessment3')
                if test3 is not None:
                    test3 =user.get('Assessment3')
                    
                    communication = int(user['Assessment3'].get('communication'))
                    technical = int(user["Assessment3"].get('technical'))
                    creativity = int(user['Assessment3'].get('creativity'))
                    projectmm = int(user['Assessment3'].get('projectmmt'))
                    timemanagement = int(user['Assessment3'].get('timemanagement'))
                    generalknowledge = int(user['Assessment3'].get('generalknowledge'))
                    interpersonal = int(user['Assessment3'].get('interpersonal'))
                    resultoriented = int(user['Assessment3'].get('resultoriented'))
                    leardership = int(user['Assessment3'].get('leardership'))
                    presentation = int(user['Assessment3'].get('presentation'))
                    entrepreneur = int(user['Assessment3'].get('entrepreneur'))
                    aptitude = int(user['Assessment3'].get('aptitude'))
                    
                    show_graph3=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    test3_skills = user["Assessment3"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p
                    
                    list_of_third_test_results=summary(test3_skills)
                    list_of_second_test_results=summary(test2_skills)
                    list_of_first_test_results=summary(test1_skills)
                    particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist,list_of_third_test_results)}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)
                    
                    def sumskills(user1,user2,user3):
                        p=[]
                        q=[]
                        r=[]
                        t1=[]
                        t2=[]
                        t3=[]            
                        for key, value in user1.items():
                            p.append(int(value))
                        t1.append(int((sum(p)/60)*5))
                        # print(t1)
                        for key, value in user2.items():
                            q.append(int(value))
                        t2.append(int((sum(q)/60)*5))
                        for key, value in user3.items():
                            r.append(int(value))
                        t3.append(int((sum(r)/60)*5) )
                        score1=t1+t2+t3           
                        return score1
                    score1= sumskills(test1_skills,test2_skills,test3_skills)
                
                    return render_template('dashboard.html',show_graph3=show_graph3,score1=json.dumps(score1),first_element=first_element,second_element=second_element,third_element=third_element,p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test3summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),test1=test,entrepreneur=entrepreneur,aptitude=aptitude,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation)
                else:
                    flash('Marks has not been allotted.', category='error')     
            if test=="Assessment4":
                test4 =user.get('Assessment4')
                if test4 is not None:

                    test4 =user.get('Assessment4')
                    
                    communication = int(user['Assessment4'].get('communication'))
                    technical = int(user["Assessment4"].get('technical'))
                    creativity = int(user['Assessment4'].get('creativity'))
                    projectmm = int(user['Assessment4'].get('projectmmt'))
                    timemanagement = int(user['Assessment4'].get('timemanagement'))
                    generalknowledge = int(user['Assessment4'].get('generalknowledge'))
                    interpersonal = int(user['Assessment4'].get('interpersonal'))
                    resultoriented = int(user['Assessment4'].get('resultoriented'))
                    leardership = int(user['Assessment4'].get('leardership'))
                    presentation = int(user['Assessment4'].get('presentation'))
                    entrepreneur = int(user['Assessment4'].get('entrepreneur'))
                    aptitude = int(user['Assessment4'].get('aptitude'))
                    show_graph4=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    test3_skills = user["Assessment3"]
                    test4_skills = user["Assessment4"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p
                    
                    list_of_fourth_test_results=summary(test4_skills)
                    
                    print(len(list_of_fourth_test_results))
                    list_of_third_test_results=summary(test3_skills)
                    list_of_second_test_results=summary(test2_skills)
                    list_of_first_test_results=summary(test1_skills)
                    particular_test4summary=int((sum(list_of_fourth_test_results)/60)*5)
                    print(particular_test4summary)
                    particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist,list_of_fourth_test_results)}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)
                    def sumskills(user1,user2,user3,user4):
                        p=[]
                        q=[]
                        r=[]
                        s=[]
                        t1=[]
                        t2=[]
                        t3=[]            
                        t4=[]            
                        for key, value in user1.items():
                            p.append(int(value))
                        t1.append(int((sum(p)/60)*5))
                        # print(t1)
                        for key, value in user2.items():
                            q.append(int(value))
                        t2.append(int((sum(q)/60)*5))
                        for key, value in user3.items():
                            r.append(int(value))
                        t3.append(int((sum(r)/60)*5) )
                        for key, value in user4.items():
                            s.append(int(value))
                        t4.append(int((sum(s)/60)*5) )
                        score1=t1+t2+t3+t4          
                        return score1
                    score1= sumskills(test1_skills,test2_skills,test3_skills,test4_skills)
                
                    return render_template('dashboard.html',show_graph4=show_graph4,first_element=first_element,second_element=second_element,third_element=third_element,score1=json.dumps(score1),p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test4summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),s=json.dumps(list_of_fourth_test_results),test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)
                else:
                    flash('Marks has not been allotted.', category='error')     
            if test=="Assessment5":
                test4 =user.get('Assessment5')
                if test4 is not None:
                    test4 =user.get('Assessment5')
                    
                    communication = int(user['Assessment5'].get('communication'))
                    technical = int(user["Assessment5"].get('technical'))
                    creativity = int(user['Assessment5'].get('creativity'))
                    projectmm = int(user['Assessment5'].get('projectmmt'))
                    timemanagement = int(user['Assessment5'].get('timemanagement'))
                    generalknowledge = int(user['Assessment5'].get('generalknowledge'))
                    interpersonal = int(user['Assessment5'].get('interpersonal'))
                    resultoriented = int(user['Assessment5'].get('resultoriented'))
                    leardership = int(user['Assessment5'].get('leardership'))
                    presentation = int(user['Assessment5'].get('presentation'))
                    entrepreneur = int(user['Assessment5'].get('entrepreneur'))
                    aptitude = int(user['Assessment5'].get('aptitude'))
                    show_graph5=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    test3_skills = user["Assessment3"]
                    test4_skills = user["Assessment4"]
                    test5_skills = user["Assessment5"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p
                    
                    list_of_fifth_test_results=summary(test5_skills)
                    list_of_fourth_test_results=summary(test4_skills)
                    list_of_third_test_results=summary(test3_skills)
                    list_of_second_test_results=summary(test2_skills)
                    list_of_first_test_results=summary(test1_skills)
                    particular_test5summary=int((sum(list_of_fifth_test_results)/60)*5)
                    particular_test4summary=int((sum(list_of_fourth_test_results)/60)*5)
                    particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist,list_of_fifth_test_results)}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)
                    def sumskills(user1,user2,user3,user4,user5):
                        p=[]
                        q=[]
                        r=[]
                        s=[]
                        t=[]
                        
                        t1=[]
                        t2=[]
                        t3=[]            
                        t4=[]            
                        t5=[]            
                                    
                        for key, value in user1.items():
                            p.append(int(value))
                        t1.append(int((sum(p)/60)*5))
                        # print(t1)
                        for key, value in user2.items():
                            q.append(int(value))
                        t2.append(int((sum(q)/60)*5))
                        for key, value in user3.items():
                            r.append(int(value))
                        t3.append(int((sum(r)/60)*5) )
                        for key, value in user4.items():
                            s.append(int(value))
                        t4.append(int((sum(s)/60)*5) )
                        for key, value in user5.items():
                            t.append(int(value))
                        t5.append(int((sum(t)/60)*5) )
                    
                        score1=t1+t2+t3+t4+t5  
                        print(score1)     
                        return score1
                    score1= sumskills(test1_skills,test2_skills,test3_skills,test4_skills,test5_skills)
                
                    return render_template('dashboard.html',show_graph5=show_graph5,first_element=first_element,second_element=second_element,third_element=third_element,score1=json.dumps(score1),p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test5summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),s=json.dumps(list_of_fourth_test_results),t=json.dumps(list_of_fifth_test_results),test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)
                else:
                    flash('Marks has not been allotted.', category='error')     
            if test=="Assessment6":
                test4 =user.get('Assessment6')
                if test4 is not None:

                    test4 =user.get('Assessment6')
                    
                    communication = int(user['Assessment6'].get('communication'))
                    technical = int(user["Assessment6"].get('technical'))
                    creativity = int(user['Assessment6'].get('creativity'))
                    projectmm = int(user['Assessment6'].get('projectmmt'))
                    timemanagement = int(user['Assessment6'].get('timemanagement'))
                    generalknowledge = int(user['Assessment6'].get('generalknowledge'))
                    interpersonal = int(user['Assessment6'].get('interpersonal'))
                    resultoriented = int(user['Assessment6'].get('resultoriented'))
                    leardership = int(user['Assessment6'].get('leardership'))
                    presentation = int(user['Assessment6'].get('presentation'))
                    entrepreneur = int(user['Assessment6'].get('entrepreneur'))
                    aptitude = int(user['Assessment6'].get('aptitude'))
                    show_graph6=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    test3_skills = user["Assessment3"]
                    test4_skills = user["Assessment4"]
                    test5_skills = user["Assessment5"]
                    test6_skills = user["Assessment6"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p
                    
                    list_of_sixth_test_results=summary(test6_skills)
                    list_of_fifth_test_results=summary(test5_skills)
                    list_of_fourth_test_results=summary(test4_skills)
                    list_of_third_test_results=summary(test3_skills)
                    list_of_second_test_results=summary(test2_skills)
                    list_of_first_test_results=summary(test1_skills)
                    particular_test6summary=int((sum(list_of_sixth_test_results)/60)*5)
                    particular_test5summary=int((sum(list_of_fifth_test_results)/60)*5)
                    particular_test4summary=int((sum(list_of_fourth_test_results)/60)*5)
                    print(particular_test6summary)
                    particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist,list_of_sixth_test_results)}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)
                    def sumskills(user1,user2,user3,user4,user5,user6):
                        p=[]
                        q=[]
                        r=[]
                        s=[]
                        t=[]
                        u=[]
                        t1=[]
                        t2=[]
                        t3=[]            
                        t4=[]            
                        t5=[]            
                        t6=[]            
                        for key, value in user1.items():
                            p.append(int(value))
                        t1.append(int((sum(p)/60)*5))
                        # print(t1)
                        for key, value in user2.items():
                            q.append(int(value))
                        t2.append(int((sum(q)/60)*5))
                        for key, value in user3.items():
                            r.append(int(value))
                        t3.append(int((sum(r)/60)*5) )
                        for key, value in user4.items():
                            s.append(int(value))
                        t4.append(int((sum(s)/60)*5) )
                        for key, value in user5.items():
                            t.append(int(value))
                        t5.append(int((sum(t)/60)*5) )
                        for key, value in user6.items():
                            u.append(int(value))
                        t6.append(int((sum(u)/60)*5) )
                        score1=t1+t2+t3+t4+t5+t6 
                        print(t6)
                            
                        return score1
                    score1= sumskills(test1_skills,test2_skills,test3_skills,test4_skills,test5_skills,test6_skills)
                    print(score1) 
                
                    return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,show_graph6=show_graph6,score1=json.dumps(score1),p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test6summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),s=json.dumps(list_of_fourth_test_results),t=json.dumps(list_of_fifth_test_results),u=json.dumps(list_of_sixth_test_results),test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)
                else:
                    flash('Marks has not been allotted.', category='error')     
            if test=="Assessment7":
                test4 =user.get('Assessment7')
                if test4 is not None:
                    test4 =user.get('Assessment7')
                    
                    communication = int(user['Assessment7'].get('communication'))
                    technical = int(user["Assessment7"].get('technical'))
                    creativity = int(user['Assessment7'].get('creativity'))
                    projectmm = int(user['Assessment7'].get('projectmmt'))
                    timemanagement = int(user['Assessment7'].get('timemanagement'))
                    generalknowledge = int(user['Assessment7'].get('generalknowledge'))
                    interpersonal = int(user['Assessment7'].get('interpersonal'))
                    resultoriented = int(user['Assessment7'].get('resultoriented'))
                    leardership = int(user['Assessment7'].get('leardership'))
                    presentation = int(user['Assessment7'].get('presentation'))
                    entrepreneur = int(user['Assessment7'].get('entrepreneur'))
                    aptitude = int(user['Assessment7'].get('aptitude'))
                    show_graph7=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    test3_skills = user["Assessment3"]
                    test4_skills = user["Assessment4"]
                    test5_skills = user["Assessment5"]
                    test6_skills = user["Assessment6"]
                    test7_skills = user["Assessment7"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p
                    
                    list_of_seventh_test_results=summary(test7_skills)
                    list_of_sixth_test_results=summary(test6_skills)
                    list_of_fifth_test_results=summary(test5_skills)
                    list_of_fourth_test_results=summary(test4_skills)
                    list_of_third_test_results=summary(test3_skills)
                    list_of_second_test_results=summary(test2_skills)
                    list_of_first_test_results=summary(test1_skills)
                    particular_test7summary=int((sum(list_of_seventh_test_results)/60)*5)
                    particular_test6summary=int((sum(list_of_sixth_test_results)/60)*5)
                    particular_test5summary=int((sum(list_of_fifth_test_results)/60)*5)
                    particular_test4summary=int((sum(list_of_fourth_test_results)/60)*5)
                    print(particular_test6summary)
                    particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist,list_of_seventh_test_results)}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)
                    def sumskills(user1,user2,user3,user4,user5,user6,user7):
                        p=[]
                        q=[]
                        r=[]
                        s=[]
                        t=[]
                        u=[]
                        v=[]
                        t1=[]
                        t2=[]
                        t3=[]            
                        t4=[]            
                        t5=[]            
                        t6=[]            
                        t7=[]            
                        for key, value in user1.items():
                            p.append(int(value))
                        t1.append(int((sum(p)/60)*5))
                        # print(t1)
                        for key, value in user2.items():
                            q.append(int(value))
                        t2.append(int((sum(q)/60)*5))
                        for key, value in user3.items():
                            r.append(int(value))
                        t3.append(int((sum(r)/60)*5) )
                        for key, value in user4.items():
                            s.append(int(value))
                        t4.append(int((sum(s)/60)*5) )
                        for key, value in user5.items():
                            t.append(int(value))
                        t5.append(int((sum(t)/60)*5) )
                        for key, value in user6.items():
                            u.append(int(value))
                        t6.append(int((sum(u)/60)*5) )
                        for key, value in user7.items():
                            v.append(int(value))
                        t7.append(int((sum(v)/60)*5) )
                        score1=t1+t2+t3+t4+t5+t6+t7
                        print(t6)
                            
                        return score1
                    score1= sumskills(test1_skills,test2_skills,test3_skills,test4_skills,test5_skills,test6_skills,test7_skills)
                    print(score1) 
                
                    return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,show_graph7=show_graph7,score1=json.dumps(score1),p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test7summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),s=json.dumps(list_of_fourth_test_results),t=json.dumps(list_of_fifth_test_results),u=json.dumps(list_of_sixth_test_results),v=json.dumps(list_of_seventh_test_results),test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)
                else:
                    flash('Marks has not been allotted.', category='error')     
            if test=="Assessment8":
                test4 =user.get('Assessment8')
                if test4 is not None:

                    test4 =user.get('Assessment8')
                    
                    communication = int(user['Assessment8'].get('communication'))
                    technical = int(user["Assessment8"].get('technical'))
                    creativity = int(user['Assessment8'].get('creativity'))
                    projectmm = int(user['Assessment8'].get('projectmmt'))
                    timemanagement = int(user['Assessment8'].get('timemanagement'))
                    generalknowledge = int(user['Assessment8'].get('generalknowledge'))
                    interpersonal = int(user['Assessment8'].get('interpersonal'))
                    resultoriented = int(user['Assessment8'].get('resultoriented'))
                    leardership = int(user['Assessment8'].get('leardership'))
                    presentation = int(user['Assessment8'].get('presentation'))
                    entrepreneur = int(user['Assessment8'].get('entrepreneur'))
                    aptitude = int(user['Assessment8'].get('aptitude'))
                    show_graph8=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    test3_skills = user["Assessment3"]
                    test4_skills = user["Assessment4"]
                    test5_skills = user["Assessment5"]
                    test6_skills = user["Assessment6"]
                    test7_skills = user["Assessment7"]
                    test8_skills = user["Assessment8"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p
                    
                    list_of_eightth_test_results=summary(test8_skills)
                    list_of_seventh_test_results=summary(test7_skills)
                    list_of_sixth_test_results=summary(test6_skills)
                    list_of_fifth_test_results=summary(test5_skills)
                    list_of_fourth_test_results=summary(test4_skills)
                    list_of_third_test_results=summary(test3_skills)
                    list_of_second_test_results=summary(test2_skills)
                    list_of_first_test_results=summary(test1_skills)
                    particular_test8summary=int((sum(list_of_eightth_test_results)/60)*5)
                    particular_test7summary=int((sum(list_of_seventh_test_results)/60)*5)
                    particular_test6summary=int((sum(list_of_sixth_test_results)/60)*5)
                    particular_test5summary=int((sum(list_of_fifth_test_results)/60)*5)
                    particular_test4summary=int((sum(list_of_fourth_test_results)/60)*5)
                    print(particular_test6summary)
                    particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist,list_of_eightth_test_results)}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)
                    def sumskills(user1,user2,user3,user4,user5,user6,user7,user8):
                        p=[]
                        q=[]
                        r=[]
                        s=[]
                        t=[]
                        u=[]
                        v=[]
                        w=[]
                        t1=[]
                        t2=[]
                        t3=[]            
                        t4=[]            
                        t5=[]            
                        t6=[]            
                        t7=[]            
                        t8=[]            
                        for key, value in user1.items():
                            p.append(int(value))
                        t1.append(int((sum(p)/60)*5))
                        # print(t1)
                        for key, value in user2.items():
                            q.append(int(value))
                        t2.append(int((sum(q)/60)*5))
                        for key, value in user3.items():
                            r.append(int(value))
                        t3.append(int((sum(r)/60)*5) )
                        for key, value in user4.items():
                            s.append(int(value))
                        t4.append(int((sum(s)/60)*5) )
                        for key, value in user5.items():
                            t.append(int(value))
                        t5.append(int((sum(t)/60)*5) )
                        for key, value in user6.items():
                            u.append(int(value))
                        t6.append(int((sum(u)/60)*5) )
                        for key, value in user7.items():
                            v.append(int(value))
                        t7.append(int((sum(v)/60)*5) )
                        for key, value in user8.items():
                            w.append(int(value))
                        t8.append(int((sum(w)/60)*5) )
                        score1=t1+t2+t3+t4+t5+t6+t7+t8
                        print(t6)
                            
                        return score1
                    score1= sumskills(test1_skills,test2_skills,test3_skills,test4_skills,test5_skills,test6_skills,test7_skills,test8_skills)
                    print(score1) 
                
                    return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,show_graph8=show_graph8,score1=json.dumps(score1),p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test8summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),s=json.dumps(list_of_fourth_test_results),t=json.dumps(list_of_fifth_test_results),u=json.dumps(list_of_sixth_test_results),v=json.dumps(list_of_seventh_test_results),w=json.dumps(list_of_eightth_test_results),test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)
                else:
                    flash('Marks has not been allotted.', category='error')     
        communication = int(user['Assessment1'].get('communication'))
        technical = int(user["Assessment1"].get('technical'))
        creativity = int(user['Assessment1'].get('creativity'))
        projectmm = int(user['Assessment1'].get('projectmmt'))
        timemanagement = int(user['Assessment1'].get('timemanagement'))
        generalknowledge = int(user['Assessment1'].get('generalknowledge'))
        interpersonal = int(user['Assessment1'].get('interpersonal'))
        resultoriented = int(user['Assessment1'].get('resultoriented'))
        leardership = int(user['Assessment1'].get('leardership'))
        presentation = int(user['Assessment1'].get('presentation'))
        entrepreneur = int(user['Assessment1'].get('entrepreneur'))
        aptitude = int(user['Assessment1'].get('aptitude'))

         
                       
        test11="Assessment1"
      


        return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,particular_testsummary=particular_testsummary,test1=test11,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)

        


        
      


        title="dashboard"

      
    elif  'teacheremail' in session :
        title="dashboard"
        usn = session.get("studentname")
        user = db.users.find_one({'usn':usn })
        username = user["personal"].get('username')
        test1_skills = user["Assessment1"]
        def summary(userr):
            p=[]
            for key, value in userr.items():
                p.append(int(value))
            return p
        test_function_summary=summary(test1_skills)  
        skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
        elements={key:value for key,value in zip(skilllist, test_function_summary)}
        # Sort the dictionary based on values (descending order)
        sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
        iterator = iter(sortedinorder.items())
        first_element=next(iterator)
        # next(iterator)  # Skip the first element
        second_element = next(iterator)
        # next(iterator) 
        third_element = next(iterator)     
        particular_testsummary=int((sum(test_function_summary)/60)*5)  
               
        if request.method=="POST":
            man= request.form.get('test')
            test =man
            print(test)
            title="dashboard"
            usn = session.get("studentname")
            user = db.users.find_one({'usn':usn })
            username = user["personal"].get('username')
            show_graph1=False
            show_graph2=False
            show_graph3=False
            show_graph4=False
            show_graph5=False
            show_graph6=False
            show_graph7=False
            show_graph8=False
            

            # print(test)
            if test=="Assessment1":
             
                communication = int(user['Assessment1'].get('communication'))
                technical = int(user["Assessment1"].get('technical'))
                creativity = int(user['Assessment1'].get('creativity'))
                projectmm = int(user['Assessment1'].get('projectmmt'))
                timemanagement = int(user['Assessment1'].get('timemanagement'))
                generalknowledge = int(user['Assessment1'].get('generalknowledge'))
                interpersonal = int(user['Assessment1'].get('interpersonal'))
                resultoriented = int(user['Assessment1'].get('resultoriented'))
                leardership = int(user['Assessment1'].get('leardership'))
                presentation = int(user['Assessment1'].get('presentation'))
                entrepreneur = int(user['Assessment1'].get('entrepreneur'))
                aptitude = int(user['Assessment1'].get('aptitude'))

                show_graph1=True
                def summary(userr):
                    p=[]
                    for key, value in userr.items():
                        p.append(int(value))
                    return p
                p=summary(test1_skills) 
                skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                elements={key:value for key,value in zip(skilllist, p)}
                # Sort the dictionary based on values (descending order)
                sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                iterator = iter(sortedinorder.items())
                first_element=next(iterator)
                # next(iterator)  # Skip the first element
                second_element = next(iterator)
                # next(iterator) 
                third_element = next(iterator)     
                
                
                particular_test1summary=int((sum(p)/60)*5)
                return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,show_graph1=show_graph1,particular_testsummary=particular_test1summary,test1=test,entrepreneur=entrepreneur,aptitude=aptitude,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation)

            if test=="Assessment2":
                test2 =user.get('Assessment2')

                if test2 is not None:
                
                    test2 =user.get('Assessment2')

                
                
                    communication = int(user['Assessment2'].get('communication'))
                    technical = int(user["Assessment2"].get('technical'))
                    creativity = int(user['Assessment2'].get('creativity'))
                    projectmm = int(user['Assessment2'].get('projectmmt'))
                    timemanagement = int(user['Assessment2'].get('timemanagement'))
                    generalknowledge = int(user['Assessment2'].get('generalknowledge'))
                    interpersonal = int(user['Assessment2'].get('interpersonal'))
                    resultoriented = int(user['Assessment2'].get('resultoriented'))
                    leardership = int(user['Assessment2'].get('leardership'))
                    presentation = int(user['Assessment2'].get('presentation'))
                    entrepreneur = int(user['Assessment2'].get('entrepreneur'))
                    aptitude = int(user['Assessment2'].get('aptitude'))
                    show_graph2=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p                
                    list_of_second_test_results=summary(test2_skills)            
                    list_of_first_test_results=summary(test1_skills)            
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist, list_of_second_test_results )}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)     
                    
                    return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,show_graph2=show_graph2,particular_testsummary=particular_test2summary,p=json.dumps(list_of_first_test_results),q=json.dumps(list_of_second_test_results),entrepreneur=entrepreneur,aptitude=aptitude,test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation)
                else:
                    flash('Marks has not been allotted.', category='error')                # redirect(url_for('views.dashboard'))
  
            if test=="Assessment3":
                test3 =user.get('Assessment3')
                if test3 is not None:
                    test3 =user.get('Assessment3')

                
                    communication = int(user['Assessment3'].get('communication'))
                    technical = int(user["Assessment3"].get('technical'))
                    creativity = int(user['Assessment3'].get('creativity'))
                    projectmm = int(user['Assessment3'].get('projectmmt'))
                    timemanagement = int(user['Assessment3'].get('timemanagement'))
                    generalknowledge = int(user['Assessment3'].get('generalknowledge'))
                    interpersonal = int(user['Assessment3'].get('interpersonal'))
                    resultoriented = int(user['Assessment3'].get('resultoriented'))
                    leardership = int(user['Assessment3'].get('leardership'))
                    presentation = int(user['Assessment3'].get('presentation'))
                    entrepreneur = int(user['Assessment3'].get('entrepreneur'))
                    aptitude = int(user['Assessment3'].get('aptitude'))
                    
                    show_graph3=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    test3_skills = user["Assessment3"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p
                    
                    list_of_third_test_results=summary(test3_skills)
                    list_of_second_test_results=summary(test2_skills)
                    list_of_first_test_results=summary(test1_skills)
                    particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist, list_of_third_test_results )}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)     
                    def sumskills(user1,user2,user3):
                        p=[]
                        q=[]
                        r=[]
                        t1=[]
                        t2=[]
                        t3=[]            
                        for key, value in user1.items():
                            p.append(int(value))
                        t1.append(int((sum(p)/60)*5))
                        # print(t1)
                        for key, value in user2.items():
                            q.append(int(value))
                        t2.append(int((sum(q)/60)*5))
                        for key, value in user3.items():
                            r.append(int(value))
                        t3.append(int((sum(r)/60)*5) )
                        score1=t1+t2+t3           
                        return score1
                    score1= sumskills(test1_skills,test2_skills,test3_skills)
                
                    return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,show_graph3=show_graph3,score1=json.dumps(score1),p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test3summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),test1=test,entrepreneur=entrepreneur,aptitude=aptitude,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation)
                if test=="Assessment4":
                    print("Iam in")
                    test4 =user.get('Assessment4')
                    if test4 is not None:
                        test4 =user.get('Assessment4')                       
                        communication = int(user['Assessment4'].get('communication'))
                        technical = int(user["Assessment4"].get('technical'))
                        creativity = int(user['Assessment4'].get('creativity'))
                        projectmm = int(user['Assessment4'].get('projectmmt'))
                        timemanagement = int(user['Assessment4'].get('timemanagement'))
                        generalknowledge = int(user['Assessment4'].get('generalknowledge'))
                        interpersonal = int(user['Assessment4'].get('interpersonal'))
                        resultoriented = int(user['Assessment4'].get('resultoriented'))
                        leardership = int(user['Assessment4'].get('leardership'))
                        presentation = int(user['Assessment4'].get('presentation'))
                        entrepreneur = int(user['Assessment4'].get('entrepreneur'))
                        aptitude = int(user['Assessment4'].get('aptitude'))
                        show_graph4=True
                        test1_skills = user["Assessment1"]
                        test2_skills = user["Assessment2"]
                        test3_skills = user["Assessment3"]
                        test4_skills = user["Assessment4"]
                        print("fyvuybibiubhibh")
                        def summary(userr):
                            p=[]
                            for key, value in userr.items():
                                p.append(int(value))
                            return p
                        
                        list_of_fourth_test_results=summary(test4_skills)
                        print(len(list_of_fourth_test_results))
                        list_of_third_test_results=summary(test3_skills)
                        list_of_second_test_results=summary(test2_skills)
                        list_of_first_test_results=summary(test1_skills)
                        particular_test4summary=int((sum(list_of_fourth_test_results)/60)*5)
                        print(particular_test4summary)
                        particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                        particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                        particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                        skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                        elements={key:value for key,value in zip(skilllist, list_of_fourth_test_results )}
                        # Sort the dictionary based on values (descending order)
                        sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                        iterator = iter(sortedinorder.items())
                        first_element=next(iterator)
                        # next(iterator)  # Skip the first element
                        second_element = next(iterator)
                        # next(iterator) 
                        third_element = next(iterator)     
                        def sumskills(user1,user2,user3,user4):
                            p=[]
                            q=[]
                            r=[]
                            s=[]
                            t1=[]
                            t2=[]
                            t3=[]            
                            t4=[]            
                            for key, value in user1.items():
                                p.append(int(value))
                            t1.append(int((sum(p)/60)*5))
                            # print(t1)
                            for key, value in user2.items():
                                q.append(int(value))
                            t2.append(int((sum(q)/60)*5))
                            for key, value in user3.items():
                                r.append(int(value))
                            t3.append(int((sum(r)/60)*5) )
                            for key, value in user4.items():
                                s.append(int(value))
                            t4.append(int((sum(s)/60)*5) )
                            score1=t1+t2+t3+t4          
                            return score1
                        score1= sumskills(test1_skills,test2_skills,test3_skills,test4_skills)
                    
                        return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,show_graph4=show_graph4,score1=json.dumps(score1),p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test4summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),s=json.dumps(list_of_fourth_test_results),test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)
                    else:

                        flash('Marks has not been allotted.', category='error')                # redirect(url_for('views.dashboard'))
            if test=="Assessment4":
                test4 =user.get('Assessment4')
                if test4 is not None:

                    test4 =user.get('Assessment4')
                    
                    communication = int(user['Assessment4'].get('communication'))
                    technical = int(user["Assessment4"].get('technical'))
                    creativity = int(user['Assessment4'].get('creativity'))
                    projectmm = int(user['Assessment4'].get('projectmmt'))
                    timemanagement = int(user['Assessment4'].get('timemanagement'))
                    generalknowledge = int(user['Assessment4'].get('generalknowledge'))
                    interpersonal = int(user['Assessment4'].get('interpersonal'))
                    resultoriented = int(user['Assessment4'].get('resultoriented'))
                    leardership = int(user['Assessment4'].get('leardership'))
                    presentation = int(user['Assessment4'].get('presentation'))
                    entrepreneur = int(user['Assessment4'].get('entrepreneur'))
                    aptitude = int(user['Assessment4'].get('aptitude'))
                    show_graph4=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    test3_skills = user["Assessment3"]
                    test4_skills = user["Assessment4"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p
                    
                    list_of_fourth_test_results=summary(test4_skills)
                    
                    print(len(list_of_fourth_test_results))
                    list_of_third_test_results=summary(test3_skills)
                    list_of_second_test_results=summary(test2_skills)
                    list_of_first_test_results=summary(test1_skills)
                    particular_test4summary=int((sum(list_of_fourth_test_results)/60)*5)
                    print(particular_test4summary)
                    particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist,list_of_fourth_test_results)}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)
                    def sumskills(user1,user2,user3,user4):
                        p=[]
                        q=[]
                        r=[]
                        s=[]
                        t1=[]
                        t2=[]
                        t3=[]            
                        t4=[]            
                        for key, value in user1.items():
                            p.append(int(value))
                        t1.append(int((sum(p)/60)*5))
                        # print(t1)
                        for key, value in user2.items():
                            q.append(int(value))
                        t2.append(int((sum(q)/60)*5))
                        for key, value in user3.items():
                            r.append(int(value))
                        t3.append(int((sum(r)/60)*5) )
                        for key, value in user4.items():
                            s.append(int(value))
                        t4.append(int((sum(s)/60)*5) )
                        score1=t1+t2+t3+t4          
                        return score1
                    score1= sumskills(test1_skills,test2_skills,test3_skills,test4_skills)
                
                    return render_template('dashboard.html',show_graph4=show_graph4,first_element=first_element,second_element=second_element,third_element=third_element,score1=json.dumps(score1),p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test4summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),s=json.dumps(list_of_fourth_test_results),test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)
                else:
                    flash('Marks has not been allotted.', category='error')     
            if test=="Assessment5":
                test4 =user.get('Assessment5')
                if test4 is not None:
                    test4 =user.get('Assessment5')
                    
                    communication = int(user['Assessment5'].get('communication'))
                    technical = int(user["Assessment5"].get('technical'))
                    creativity = int(user['Assessment5'].get('creativity'))
                    projectmm = int(user['Assessment5'].get('projectmmt'))
                    timemanagement = int(user['Assessment5'].get('timemanagement'))
                    generalknowledge = int(user['Assessment5'].get('generalknowledge'))
                    interpersonal = int(user['Assessment5'].get('interpersonal'))
                    resultoriented = int(user['Assessment5'].get('resultoriented'))
                    leardership = int(user['Assessment5'].get('leardership'))
                    presentation = int(user['Assessment5'].get('presentation'))
                    entrepreneur = int(user['Assessment5'].get('entrepreneur'))
                    aptitude = int(user['Assessment5'].get('aptitude'))
                    show_graph5=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    test3_skills = user["Assessment3"]
                    test4_skills = user["Assessment4"]
                    test5_skills = user["Assessment5"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p
                    
                    list_of_fifth_test_results=summary(test5_skills)
                    list_of_fourth_test_results=summary(test4_skills)
                    list_of_third_test_results=summary(test3_skills)
                    list_of_second_test_results=summary(test2_skills)
                    list_of_first_test_results=summary(test1_skills)
                    particular_test5summary=int((sum(list_of_fifth_test_results)/60)*5)
                    particular_test4summary=int((sum(list_of_fourth_test_results)/60)*5)
                    particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist, list_of_fifth_test_results )}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)     
                    def sumskills(user1,user2,user3,user4,user5):
                        p=[]
                        q=[]
                        r=[]
                        s=[]
                        t=[]
                        
                        t1=[]
                        t2=[]
                        t3=[]            
                        t4=[]            
                        t5=[]            
                                    
                        for key, value in user1.items():
                            p.append(int(value))
                        t1.append(int((sum(p)/60)*5))
                        # print(t1)
                        for key, value in user2.items():
                            q.append(int(value))
                        t2.append(int((sum(q)/60)*5))
                        for key, value in user3.items():
                            r.append(int(value))
                        t3.append(int((sum(r)/60)*5) )
                        for key, value in user4.items():
                            s.append(int(value))
                        t4.append(int((sum(s)/60)*5) )
                        for key, value in user5.items():
                            t.append(int(value))
                        t5.append(int((sum(t)/60)*5) )
                    
                        score1=t1+t2+t3+t4+t5  
                        print(score1)     
                        return score1
                    score1= sumskills(test1_skills,test2_skills,test3_skills,test4_skills,test5_skills)
                
                    return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,show_graph5=show_graph5,score1=json.dumps(score1),p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test5summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),s=json.dumps(list_of_fourth_test_results),t=json.dumps(list_of_fifth_test_results),test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)
                else:
                    flash('Marks has not been allotted.', category='error')                # redirect(url_for('views.dashboard'))
  
            if test=="Assessment6":
                test4 =user.get('Assessment6')
                if test4 is not None:
                    test4 =user.get('Assessment6')
                
                    communication = int(user['Assessment6'].get('communication'))
                    technical = int(user["Assessment6"].get('technical'))
                    creativity = int(user['Assessment6'].get('creativity'))
                    projectmm = int(user['Assessment6'].get('projectmmt'))
                    timemanagement = int(user['Assessment6'].get('timemanagement'))
                    generalknowledge = int(user['Assessment6'].get('generalknowledge'))
                    interpersonal = int(user['Assessment6'].get('interpersonal'))
                    resultoriented = int(user['Assessment6'].get('resultoriented'))
                    leardership = int(user['Assessment6'].get('leardership'))
                    presentation = int(user['Assessment6'].get('presentation'))
                    entrepreneur = int(user['Assessment6'].get('entrepreneur'))
                    aptitude = int(user['Assessment6'].get('aptitude'))
                    show_graph6=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    test3_skills = user["Assessment3"]
                    test4_skills = user["Assessment4"]
                    test5_skills = user["Assessment5"]
                    test6_skills = user["Assessment6"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p
                    
                    list_of_sixth_test_results=summary(test6_skills)
                    list_of_fifth_test_results=summary(test5_skills)
                    list_of_fourth_test_results=summary(test4_skills)
                    list_of_third_test_results=summary(test3_skills)
                    list_of_second_test_results=summary(test2_skills)
                    list_of_first_test_results=summary(test1_skills)
                    particular_test6summary=int((sum(list_of_sixth_test_results)/60)*5)
                    particular_test5summary=int((sum(list_of_fifth_test_results)/60)*5)
                    particular_test4summary=int((sum(list_of_fourth_test_results)/60)*5)
                    print(particular_test6summary)
                    particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist, list_of_sixth_test_results )}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)     
                    def sumskills(user1,user2,user3,user4,user5,user6):
                        p=[]
                        q=[]
                        r=[]
                        s=[]
                        t=[]
                        u=[]
                        t1=[]
                        t2=[]
                        t3=[]            
                        t4=[]            
                        t5=[]            
                        t6=[]            
                        for key, value in user1.items():
                            p.append(int(value))
                        t1.append(int((sum(p)/60)*5))
                        # print(t1)
                        for key, value in user2.items():
                            q.append(int(value))
                        t2.append(int((sum(q)/60)*5))
                        for key, value in user3.items():
                            r.append(int(value))
                        t3.append(int((sum(r)/60)*5) )
                        for key, value in user4.items():
                            s.append(int(value))
                        t4.append(int((sum(s)/60)*5) )
                        for key, value in user5.items():
                            t.append(int(value))
                        t5.append(int((sum(t)/60)*5) )
                        for key, value in user6.items():
                            u.append(int(value))
                        t6.append(int((sum(u)/60)*5) )
                        score1=t1+t2+t3+t4+t5+t6 
                        print(t6)
                            
                        return score1
                    score1= sumskills(test1_skills,test2_skills,test3_skills,test4_skills,test5_skills,test6_skills)
                    print(score1) 
                
                    return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,show_graph6=show_graph6,score1=json.dumps(score1),p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test6summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),s=json.dumps(list_of_fourth_test_results),t=json.dumps(list_of_fifth_test_results),u=json.dumps(list_of_sixth_test_results),test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)
                else:
                    flash('Marks has not been allotted.', category='error')                # redirect(url_for('views.dashboard'))
  
                if test=="Assessment7":
                    test4 =user.get('Assessment7')
                    if test4 is not None:
                        test4 =user.get('Assessment7')
                        
                        communication = int(user['Assessment7'].get('communication'))
                        technical = int(user["Assessment7"].get('technical'))
                        creativity = int(user['Assessment7'].get('creativity'))
                        projectmm = int(user['Assessment7'].get('projectmmt'))
                        timemanagement = int(user['Assessment7'].get('timemanagement'))
                        generalknowledge = int(user['Assessment7'].get('generalknowledge'))
                        interpersonal = int(user['Assessment7'].get('interpersonal'))
                        resultoriented = int(user['Assessment7'].get('resultoriented'))
                        leardership = int(user['Assessment7'].get('leardership'))
                        presentation = int(user['Assessment7'].get('presentation'))
                        entrepreneur = int(user['Assessment7'].get('entrepreneur'))
                        aptitude = int(user['Assessment7'].get('aptitude'))
                        show_graph7=True
                        test1_skills = user["Assessment1"]
                        test2_skills = user["Assessment2"]
                        test3_skills = user["Assessment3"]
                        test4_skills = user["Assessment4"]
                        test5_skills = user["Assessment5"]
                        test6_skills = user["Assessment6"]
                        test7_skills = user["Assessment7"]
                        def summary(userr):
                            p=[]
                            for key, value in userr.items():
                                p.append(int(value))
                            return p
                        
                        list_of_seventh_test_results=summary(test7_skills)
                        list_of_sixth_test_results=summary(test6_skills)
                        list_of_fifth_test_results=summary(test5_skills)
                        list_of_fourth_test_results=summary(test4_skills)
                        list_of_third_test_results=summary(test3_skills)
                        list_of_second_test_results=summary(test2_skills)
                        list_of_first_test_results=summary(test1_skills)
                        particular_test7summary=int((sum(list_of_seventh_test_results)/60)*5)
                        particular_test6summary=int((sum(list_of_sixth_test_results)/60)*5)
                        particular_test5summary=int((sum(list_of_fifth_test_results)/60)*5)
                        particular_test4summary=int((sum(list_of_fourth_test_results)/60)*5)
                        print(particular_test6summary)
                        particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                        particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                        particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                        skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                        elements={key:value for key,value in zip(skilllist, list_of_seventh_test_results )}
                        # Sort the dictionary based on values (descending order)
                        sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                        iterator = iter(sortedinorder.items())
                        first_element=next(iterator)
                        # next(iterator)  # Skip the first element
                        second_element = next(iterator)
                        # next(iterator) 
                        third_element = next(iterator)     
                        def sumskills(user1,user2,user3,user4,user5,user6,user7):
                            p=[]
                            q=[]
                            r=[]
                            s=[]
                            t=[]
                            u=[]
                            v=[]
                            t1=[]
                            t2=[]
                            t3=[]            
                            t4=[]            
                            t5=[]            
                            t6=[]            
                            t7=[]            
                            for key, value in user1.items():
                                p.append(int(value))
                            t1.append(int((sum(p)/60)*5))
                            # print(t1)
                            for key, value in user2.items():
                                q.append(int(value))
                            t2.append(int((sum(q)/60)*5))
                            for key, value in user3.items():
                                r.append(int(value))
                            t3.append(int((sum(r)/60)*5) )
                            for key, value in user4.items():
                                s.append(int(value))
                            t4.append(int((sum(s)/60)*5) )
                            for key, value in user5.items():
                                t.append(int(value))
                            t5.append(int((sum(t)/60)*5) )
                            for key, value in user6.items():
                                u.append(int(value))
                            t6.append(int((sum(u)/60)*5) )
                            for key, value in user7.items():
                                v.append(int(value))
                            t7.append(int((sum(v)/60)*5) )
                            score1=t1+t2+t3+t4+t5+t6+t7
                            print(t6)
                                
                            return score1
                        score1= sumskills(test1_skills,test2_skills,test3_skills,test4_skills,test5_skills,test6_skills,test7_skills)
                        print(score1) 
                    
                        return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,show_graph7=show_graph7,score1=json.dumps(score1),p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test7summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),s=json.dumps(list_of_fourth_test_results),t=json.dumps(list_of_fifth_test_results),u=json.dumps(list_of_sixth_test_results),v=json.dumps(list_of_seventh_test_results),test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)
                    else:
                        flash('Marks has not been allotted.', category='error')                # redirect(url_for('views.dashboard'))
            if test=="Assessment7":
                test4 =user.get('Assessment7')
                if test4 is not None:
                    test4 =user.get('Assessment7')
                    
                    communication = int(user['Assessment7'].get('communication'))
                    technical = int(user["Assessment7"].get('technical'))
                    creativity = int(user['Assessment7'].get('creativity'))
                    projectmm = int(user['Assessment7'].get('projectmmt'))
                    timemanagement = int(user['Assessment7'].get('timemanagement'))
                    generalknowledge = int(user['Assessment7'].get('generalknowledge'))
                    interpersonal = int(user['Assessment7'].get('interpersonal'))
                    resultoriented = int(user['Assessment7'].get('resultoriented'))
                    leardership = int(user['Assessment7'].get('leardership'))
                    presentation = int(user['Assessment7'].get('presentation'))
                    entrepreneur = int(user['Assessment7'].get('entrepreneur'))
                    aptitude = int(user['Assessment7'].get('aptitude'))
                    show_graph7=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    test3_skills = user["Assessment3"]
                    test4_skills = user["Assessment4"]
                    test5_skills = user["Assessment5"]
                    test6_skills = user["Assessment6"]
                    test7_skills = user["Assessment7"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p
                    
                    list_of_seventh_test_results=summary(test7_skills)
                    list_of_sixth_test_results=summary(test6_skills)
                    list_of_fifth_test_results=summary(test5_skills)
                    list_of_fourth_test_results=summary(test4_skills)
                    list_of_third_test_results=summary(test3_skills)
                    list_of_second_test_results=summary(test2_skills)
                    list_of_first_test_results=summary(test1_skills)
                    particular_test7summary=int((sum(list_of_seventh_test_results)/60)*5)
                    particular_test6summary=int((sum(list_of_sixth_test_results)/60)*5)
                    particular_test5summary=int((sum(list_of_fifth_test_results)/60)*5)
                    particular_test4summary=int((sum(list_of_fourth_test_results)/60)*5)
                    print(particular_test6summary)
                    particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist,list_of_seventh_test_results)}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)
                    def sumskills(user1,user2,user3,user4,user5,user6,user7):
                        p=[]
                        q=[]
                        r=[]
                        s=[]
                        t=[]
                        u=[]
                        v=[]
                        t1=[]
                        t2=[]
                        t3=[]            
                        t4=[]            
                        t5=[]            
                        t6=[]            
                        t7=[]            
                        for key, value in user1.items():
                            p.append(int(value))
                        t1.append(int((sum(p)/60)*5))
                        # print(t1)
                        for key, value in user2.items():
                            q.append(int(value))
                        t2.append(int((sum(q)/60)*5))
                        for key, value in user3.items():
                            r.append(int(value))
                        t3.append(int((sum(r)/60)*5) )
                        for key, value in user4.items():
                            s.append(int(value))
                        t4.append(int((sum(s)/60)*5) )
                        for key, value in user5.items():
                            t.append(int(value))
                        t5.append(int((sum(t)/60)*5) )
                        for key, value in user6.items():
                            u.append(int(value))
                        t6.append(int((sum(u)/60)*5) )
                        for key, value in user7.items():
                            v.append(int(value))
                        t7.append(int((sum(v)/60)*5) )
                        score1=t1+t2+t3+t4+t5+t6+t7
                        print(t6)
                            
                        return score1
                    score1= sumskills(test1_skills,test2_skills,test3_skills,test4_skills,test5_skills,test6_skills,test7_skills)
                    print(score1) 
                
                    return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,show_graph7=show_graph7,score1=json.dumps(score1),p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test7summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),s=json.dumps(list_of_fourth_test_results),t=json.dumps(list_of_fifth_test_results),u=json.dumps(list_of_sixth_test_results),v=json.dumps(list_of_seventh_test_results),test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)
                else:
                    flash('Marks has not been allotted.', category='error')      
            if test=="Assessment8":
                test4 =user.get('Assessment8')
                if test4 is not None:

                    test4 =user.get('Assessment8')
                    
                    communication = int(user['Assessment8'].get('communication'))
                    technical = int(user["Assessment8"].get('technical'))
                    creativity = int(user['Assessment8'].get('creativity'))
                    projectmm = int(user['Assessment8'].get('projectmmt'))
                    timemanagement = int(user['Assessment8'].get('timemanagement'))
                    generalknowledge = int(user['Assessment8'].get('generalknowledge'))
                    interpersonal = int(user['Assessment8'].get('interpersonal'))
                    resultoriented = int(user['Assessment8'].get('resultoriented'))
                    leardership = int(user['Assessment8'].get('leardership'))
                    presentation = int(user['Assessment8'].get('presentation'))
                    entrepreneur = int(user['Assessment8'].get('entrepreneur'))
                    aptitude = int(user['Assessment8'].get('aptitude'))
                    show_graph8=True
                    test1_skills = user["Assessment1"]
                    test2_skills = user["Assessment2"]
                    test3_skills = user["Assessment3"]
                    test4_skills = user["Assessment4"]
                    test5_skills = user["Assessment5"]
                    test6_skills = user["Assessment6"]
                    test7_skills = user["Assessment7"]
                    test8_skills = user["Assessment8"]
                    def summary(userr):
                        p=[]
                        for key, value in userr.items():
                            p.append(int(value))
                        return p
                    
                    list_of_eightth_test_results=summary(test8_skills)
                    list_of_seventh_test_results=summary(test7_skills)
                    list_of_sixth_test_results=summary(test6_skills)
                    list_of_fifth_test_results=summary(test5_skills)
                    list_of_fourth_test_results=summary(test4_skills)
                    list_of_third_test_results=summary(test3_skills)
                    list_of_second_test_results=summary(test2_skills)
                    list_of_first_test_results=summary(test1_skills)
                    particular_test8summary=int((sum(list_of_eightth_test_results)/60)*5)
                    particular_test7summary=int((sum(list_of_seventh_test_results)/60)*5)
                    particular_test6summary=int((sum(list_of_sixth_test_results)/60)*5)
                    particular_test5summary=int((sum(list_of_fifth_test_results)/60)*5)
                    particular_test4summary=int((sum(list_of_fourth_test_results)/60)*5)
                    print(particular_test6summary)
                    particular_test3summary=int((sum(list_of_third_test_results)/60)*5)
                    particular_test2summary=int((sum(list_of_second_test_results)/60)*5)
                    particular_test1summary=int((sum(list_of_first_test_results)/60)*5)
                    skilllist=[ 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','aptitude','entrepreneur']
                    elements={key:value for key,value in zip(skilllist, list_of_eightth_test_results )}
                    # Sort the dictionary based on values (descending order)
                    sortedinorder = dict(sorted(elements.items(), key=lambda item: item[1], reverse=True))
                    iterator = iter(sortedinorder.items())
                    first_element=next(iterator)
                    # next(iterator)  # Skip the first element
                    second_element = next(iterator)
                    # next(iterator) 
                    third_element = next(iterator)     
                    def sumskills(user1,user2,user3,user4,user5,user6,user7,user8):
                        p=[]
                        q=[]
                        r=[]
                        s=[]
                        t=[]
                        u=[]
                        v=[]
                        w=[]
                        t1=[]
                        t2=[]
                        t3=[]            
                        t4=[]            
                        t5=[]            
                        t6=[]            
                        t7=[]            
                        t8=[]            
                        for key, value in user1.items():
                            p.append(int(value))
                        t1.append(int((sum(p)/60)*5))
                        # print(t1)
                        for key, value in user2.items():
                            q.append(int(value))
                        t2.append(int((sum(q)/60)*5))
                        for key, value in user3.items():
                            r.append(int(value))
                        t3.append(int((sum(r)/60)*5) )
                        for key, value in user4.items():
                            s.append(int(value))
                        t4.append(int((sum(s)/60)*5) )
                        for key, value in user5.items():
                            t.append(int(value))
                        t5.append(int((sum(t)/60)*5) )
                        for key, value in user6.items():
                            u.append(int(value))
                        t6.append(int((sum(u)/60)*5) )
                        for key, value in user7.items():
                            v.append(int(value))
                        t7.append(int((sum(v)/60)*5) )
                        for key, value in user8.items():
                            w.append(int(value))
                        t8.append(int((sum(w)/60)*5) )
                        score1=t1+t2+t3+t4+t5+t6+t7+t8
                        print(t6)
                            
                        return score1
                    score1= sumskills(test1_skills,test2_skills,test3_skills,test4_skills,test5_skills,test6_skills,test7_skills,test8_skills)
                    print(score1) 
                
                    return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,show_graph8=show_graph8,score1=json.dumps(score1),p=json.dumps(list_of_first_test_results),particular_testsummary=particular_test8summary,q=json.dumps(list_of_second_test_results),r=json.dumps(list_of_third_test_results),s=json.dumps(list_of_fourth_test_results),t=json.dumps(list_of_fifth_test_results),u=json.dumps(list_of_sixth_test_results),v=json.dumps(list_of_seventh_test_results),w=json.dumps(list_of_eightth_test_results),test1=test,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)
                
                else:
                    
                    flash('Marks has not been allotted.', category='error')                # redirect(url_for('views.dashboard'))
  
        
        

        communication = int(user['Assessment1'].get('communication'))
        technical = int(user["Assessment1"].get('technical'))
        creativity = int(user['Assessment1'].get('creativity'))
        projectmm = int(user['Assessment1'].get('projectmmt'))
        timemanagement = int(user['Assessment1'].get('timemanagement'))
        generalknowledge = int(user['Assessment1'].get('generalknowledge'))
        interpersonal = int(user['Assessment1'].get('interpersonal'))
        resultoriented = int(user['Assessment1'].get('resultoriented'))
        leardership = int(user['Assessment1'].get('leardership'))
        presentation = int(user['Assessment1'].get('presentation'))
        entrepreneur = int(user['Assessment1'].get('entrepreneur'))
        aptitude = int(user['Assessment1'].get('aptitude'))

         
                       
        test11="Assessment1"
      


        return render_template('dashboard.html',first_element=first_element,second_element=second_element,third_element=third_element,particular_testsummary=particular_testsummary,test1=test11,username=username,title=title,creativity=creativity,communication=communication,technical=technical,projectmm=projectmm,timemanagement=timemanagement,generalknowledge=generalknowledge,interpersonal=interpersonal,resultoriented=resultoriented,leardership=leardership,presentation=presentation,entrepreneur=entrepreneur,aptitude=aptitude)

        # return render_template('dashboard.html')
    return redirect(url_for('auth.login'))

@views.route("/search",methods=['GET','POST'])
def search():
    if 'teacheremail' in session:

        username = request.form.get('username')
        usn = request.form.get('usn')

        if request.method == 'POST':
            college = request.form.get('college')
            username = request.form.get('username')
            usn = request.form.get('usn')
        
            accademicyear = request.form.get('accademicyear')
            user = db.users.find_one({'usn':usn })
            # user = db.users.find_one({'email':name })
            # studname = user['username']

            # print(email)
            

            

            if user :
            
                session["studentname"] = usn
                

                user = db.users.find_one({'usn':usn })
                print(usn)
                personal = user["personal"]
                print(personal)

                if personal is None:
                    flash(' notify students to login to application', category='error')
                    return redirect(url_for('views.search'))

                # bro = session.get['studentname']
                else:
                    flash('User found Kindly enter the marks  ', category='success')
                    return redirect(url_for('views.score'))
                # return render_template('score.html')
            else:
                flash('User not found ', category='error')
                return redirect(url_for('views.search'))
                # return render_template('dashboard.html',title=title,name=name)   
        return render_template('search.html',usn=usn,username=username)
    return redirect(url_for('auth.login'))


     
@views.route('/userdata',methods=['GET','POST'])    
def userdata():
    if  "studentemail" in session:
        if request.method == "POST":
            usn = session.get("studentemail")

            username= request.form.get('username')
            email = request.form.get('email')
            phone = request.form.get('phone')
            course = request.form.get('course')
            college = request.form.get('college')
            city = request.form.get('city')
            numberofcertifications= request.form.get('numberofcertifications')       
            technicalskills= request.form.get('technicalskills')
            softskills= request.form.get('softskills')
            tenthboard= request.form.get('tenthboard')
            tenthmarks= request.form.get('tenthmarks')
            twelvethboard= request.form.get('twelvethboard')
            twelvethmarks= request.form.get('twelvethmarks')
            ugaverage= request.form.get('ugaverage')
            linkdin= request.form.get('linkdin')
            github= request.form.get('github')
            skill1= request.form.get('skill1')
            skill2= request.form.get('skill2')
            skill3= request.form.get('skill3')
            skill4= request.form.get('skill4')
            skill5= request.form.get('skill5')
            skill6= request.form.get('skill6')
            skill7= request.form.get('skill7')
            skill8= request.form.get('skill8')
            skill9= request.form.get('skill9')
            skill10= request.form.get('skill10')
            certification1= request.form.get('certification1')
            certification2= request.form.get('certification2')
            certification3= request.form.get('certification3')
            certification4= request.form.get('certification4')
            certification5= request.form.get('certification5')
            certification6= request.form.get('certification6')
            certification7= request.form.get('certification7')
            certification8= request.form.get('certification8')
            certification9= request.form.get('certification9')
            certification10= request.form.get('certification10')
            internshipexperience= request.form.get('internshipexperience')

            hiredcontent= request.form.get('hiredcontent')
           

            try:
                db.users.update_one(
                        {"usn": usn},
                        {
                            "$set": {
                                "personal": {
                                    "username": username,
                                    "email": email,
                                    "phone": phone,
                                    "course": course,
                                    "city": city,
                                    "college":college,
                                    "numberofcertifications":numberofcertifications,
                                    "technicalskills":technicalskills,
                                    "softskills":softskills,
                                    "tenthboard":tenthboard,
                                    "tenthmarks":tenthmarks,
                                    "twelvethboard":twelvethboard,
                                    "twelvethmarks":twelvethmarks,
                                    "ugaverage":ugaverage,
                                    "linkdin": linkdin,
                                    "github": github,
                                    "skill1":skill1,
                                    "skill2":skill2,
                                    "skill3":skill3,
                                    "skill4":skill4,
                                    "skill5":skill5,
                                    "skill5":skill5,
                                    "skill6":skill6,
                                    "skill7":skill7,
                                    "skill8":skill8,
                                    "skill9":skill9,
                                    "skill10":skill10,
                                    "certification1":certification1,
                                    "certification2":certification2,
                                    "certification3":certification3,
                                    "certification4":certification4,
                                    "certification5":certification5,
                                    "certification6":certification6,
                                    "certification7":certification7,
                                    "certification8":certification8,
                                    "certification9":certification9,
                                    "certification10":certification10,
                                    "internshipexperience":internshipexperience,
                                    "hiredcontent":hiredcontent,
                                }
                            }
                        }
                    )
                db.users.update_one(
                        {"usn": usn},
                        {
                            "$set": {
                                "certificate1": {
                                    
                                    
                                    "certification1":certification1,
                                    "certification2":certification2,
                                    "certification3":certification3,
                                    "certification4":certification4,
                                    "certification5":certification5,
                                    
                                }
                            }
                        }
                    )
                db.users.update_one(
                        {"usn": usn},
                        {
                            "$set": {
                                "certificate2": {
                                    
                                    
                                    "certification6":certification6,
                                    "certification7":certification7,
                                    "certification8":certification8,
                                    "certification9":certification9,
                                    "certification10":certification10,
                                    
                                }
                            }
                        }
                    )
                db.users.update_one(
                        {"usn": usn},
                        {
                            "$set": {
                                "skills1": {
                                    
                                    
                                     "skill1":skill1,
                                    "skill2":skill2,
                                    "skill3":skill3,
                                    "skill4":skill4,
                                    "skill5":skill5,
                                    
                                }
                            }
                        }
                    )
                db.users.update_one(
                        {"usn": usn},
                        {
                            "$set": {
                                "skills2": {
                                    
                                    
                                    "skill6":skill6,
                                    "skill7":skill7,
                                    "skill8":skill8,
                                    "skill9":skill9,
                                    "skill10":skill10,
                                    
                                }
                            }
                        }
                    )
                flash('details have been entered please login again', category='success')
                return redirect(url_for('auth.login'))
            except PyMongoError as e:
                flash(f'Error: {str(e)}', category='error')
                    # Handle the error accordingly, such as logging it or displaying an error message to the user
                return redirect(url_for('views.marksenter'))
        # usn = session.get('studentemail')
        # user = db.users.find_one({'usn':usn })
        # username = user["personal"].get('username')
        # email = user["personal"].get('email')
        # college = user["personal"].get('college')
        # sem = user["personal"].get('sem')
        # linkdin = user["personal"].get('linkdin')
        # github = user["personal"].get('github')
        # project1 = user["personal"].get('project-1')
        # intership1 = user["personal"].get('internship-1')
        # interst1 = user["personal"].get('interst-1')
        # title="Account"

                        
                    

        return render_template('userdata.html')
    return redirect(url_for('auth.login'))

     

# @views.route('profileskills')
# def profileskills():
#     if request.method == "POST":
#         num_inputs = int(request.form.get("num_inputs", 0))
#     else:
#         num_inputs = 0
    
#     return render_template("index.html", num_inputs=num_inputs)



@views.route('/downloadi_csv')
def downloadi_csv():
    # Connect to MongoDB

    data = db.users.find()

    # Create a new workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write headers with formatting
    sheet.append(['Serial Number', 'usn', 'Username', 'tests', 'Communication', 'Technical', 'creativity', 'projectmanagement', 'timemanagement', 'genearl knowledge', 'interpersonal', 'resultoriented', 'leadership', 'Presentation','entrepreneur','aptitude'])
    header_row = sheet[1]
    
    color_mapping = {
        None: "FF0000",  # Red
        1: "00FF00",  # Green
        2: "FFA500",  # Orange
        3: "FFFF00"  # Yellow
    }
    # Write data from MongoDB
    serial_number = 1
    for item in data:
        usn = item['usn']
        username = item.get('personal', {}).get('username', '')
        test1 = item.get('Assessment1', {})
        communication = test1.get('communication', '')
        technical = test1.get('technical', '')
        creativity = test1.get('creativity', '')
        projectmanagement = test1.get('projectmmt', '')
        timemangement = test1.get('timemanagement', '')
        generalknowledge = test1.get('generalknowledge', '')
        interpersonal = test1.get('interpersonal', '')
        resultoriented = test1.get('resultoriented', '')
        leadership = test1.get('leardership', '')
        presentation = test1.get('presentation', '')
        aptitude = test1.get('aptitude', '')
        entrepreneur = test1.get('entrepreneur', '')

        test2 = item.get('Assessment2', {})
        communications = test2.get('communication', '')
        technicals = test2.get('technical', '')
        creativitys = test2.get('creativity', '')
        projectmanagements = test2.get('projectmmt', '')
        timemangements = test2.get('timemanagement', '')
        generalknowledges = test2.get('generalknowledge', '')
        interpersonals = test2.get('interpersonal', '')
        resultorienteds = test2.get('resultoriented', '')
        leaderships = test2.get('leardership', '')
        presentations = test2.get('presentation', '')
        aptitudes = test2.get('aptitude', '')
        entrepreneurs = test2.get('entrepreneur', '')
        # print(creativitys)
        # print(type(creativitys))
        test3 = item.get('Assessment3', {})
        communicationss = test3.get('communication', '')
        technicalss = test3.get('technical', '')
        creativityss = test3.get('creativity', '')
        projectmanagementss = test3.get('projectmmt', '')
        timemangementss = test3.get('timemanagement', '')
        generalknowledgess = test3.get('generalknowledge', '')
        interpersonalss = test3.get('interpersonal', '')
        resultorientedss = test3.get('resultoriented', '')
        leadershipss = test3.get('leardership', '')
        presentationss = test3.get('presentation', '')
        aptitudess = test3.get('aptitude', '')
        entrepreneurss = test3.get('entrepreneur', '')

        test4 = item.get('Assessment4', {})
        communicationss4 = test4.get('communication', '')
        technicalss4 = test4.get('technical', '')
        creativityss4 = test4.get('creativity', '')
        projectmanagementss4 = test4.get('projectmmt', '')
        timemangementss4 = test4.get('timemanagement', '')
        generalknowledgess4 = test4.get('generalknowledge', '')
        interpersonalss4 = test4.get('interpersonal', '')
        resultorientedss4 = test4.get('resultoriented', '')
        leadershipss4 = test4.get('leardership', '')
        presentationss4 = test4.get('presentation', '')
        aptitudess4 = test4.get('aptitude', '')
        entrepreneuress4 = test4.get('entrepreneur', '')
       
        test5 = item.get('Assessment5', {})
        communicationss5 = test5.get('communication', '')
        technicalss5 = test5.get('technical', '')
        creativityss5 = test5.get('creativity', '')
        projectmanagementss5 = test5.get('projectmmt', '')
        timemangementss5 = test5.get('timemanagement', '')
        generalknowledgess5 = test5.get('generalknowledge', '')
        interpersonalss5 = test5.get('interpersonal', '')
        resultorientedss5 = test5.get('resultoriented', '')
        leadershipss5 = test5.get('leardership', '')
        presentationss5 = test5.get('presentation', '')
        aptitudess5 = test5.get('aptitude', '')
        entrepreneuress5 = test5.get('entrepreneur', '')

        test6 = item.get('Assessment6', {})
        communicationss6 = test6.get('communication', '')
        technicalss6 = test6.get('technical', '')
        creativityss6 = test6.get('creativity', '')
        projectmanagementss6 = test6.get('projectmmt', '')
        timemangementss6 = test6.get('timemanagement', '')
        generalknowledgess6 = test6.get('generalknowledge', '')
        interpersonalss6 = test6.get('interpersonal', '')
        resultorientedss6 = test6.get('resultoriented', '')
        leadershipss6 = test6.get('leardership', '')
        presentationss6 = test6.get('presentation', '')
        aptitudes6 = test6.get('aptitude', '')
        entrepreneurs6 = test6.get('entrepreneur', '')
        
        test7 = item.get('Assessment7', {})
        communicationss7 = test7.get('communication', '')
        technicalss7 = test7.get('technical', '')
        creativityss7 = test7.get('creativity', '')
        projectmanagementss7 = test7.get('projectmmt', '')
        timemangementss7 = test7.get('timemanagement', '')
        generalknowledgess7 = test7.get('generalknowledge', '')
        interpersonalss7 = test7.get('interpersonal', '')
        resultorientedss7 = test7.get('resultoriented', '')
        leadershipss7 = test7.get('leardership', '')
        presentationss7 = test7.get('presentation', '')
        aptitudes7 = test7.get('aptitude', '')
        entrepreneurs7 = test7.get('entrepreneur', '')

        test8 = item.get('Assessment8', {})
        communicationss8 = test8.get('communication', '')
        technicalss8 = test8.get('technical', '')
        creativityss8 = test8.get('creativity', '')
        projectmanagementss8 = test8.get('projectmmt', '')
        timemangementss8 = test8.get('timemanagement', '')
        generalknowledgess8 = test8.get('generalknowledge', '')
        interpersonalss8 = test8.get('interpersonal', '')
        resultorientedss8 = test8.get('resultoriented', '')
        leadershipss8 = test8.get('leardership', '')
        presentationss8 = test8.get('presentation', '')
        aptitudes8 = test8.get('aptitude', '')
        entrepreneurs8 = test8.get('entrepreneur', '')
        


        sheet.append([serial_number, usn, username, 'Assesment1', communication, technical, creativity, projectmanagement, timemangement, generalknowledge, interpersonal, resultoriented, leadership, presentation,entrepreneur,aptitude])
        sheet.append(['',  usn, username, 'Assesment2', communications, technicals, creativitys, projectmanagements, timemangements, generalknowledges, interpersonals, resultorienteds, leaderships, presentations,entrepreneurs,aptitudes])
        sheet.append(['', usn, username, 'Assesment3', communicationss, technicalss, creativityss, projectmanagementss, timemangementss, generalknowledgess, interpersonalss, resultorientedss, leadershipss, presentationss,entrepreneurss,aptitudess])
        sheet.append(['', usn, username, 'Assesment4', communicationss4, technicalss4, creativityss4, projectmanagementss4, timemangementss4, generalknowledgess4, interpersonalss4, resultorientedss4, leadershipss4, presentationss4,entrepreneuress4,aptitudess4])
        sheet.append(['', usn, username, 'Assesment5', communicationss5, technicalss5, creativityss5, projectmanagementss5, timemangementss5, generalknowledgess5, interpersonalss5, resultorientedss5, leadershipss5, presentationss5,entrepreneuress5,aptitudess5])
        sheet.append(['', usn, username, 'Assesment7', communicationss6, technicalss6, creativityss6, projectmanagementss6, timemangementss6, generalknowledgess6, interpersonalss6, resultorientedss6, leadershipss6, presentationss6,entrepreneurs6,aptitudes6])
        sheet.append(['', usn, username, 'Assesment6', communicationss7, technicalss7, creativityss7, projectmanagementss7, timemangementss7, generalknowledgess7, interpersonalss7, resultorientedss7, leadershipss7, presentationss7,entrepreneurs7,aptitudes7])
        sheet.append(['', usn, username, 'Assesment8', communicationss8, technicalss8, creativityss8, projectmanagementss8, timemangementss8, generalknowledgess8, interpersonalss8, resultorientedss8, leadershipss8, presentationss8,entrepreneurs8,aptitudes8])
        # sheet.append([''])
    
        serial_number += 1
       # Apply background color to cells with values 1, 2, 3, 4, 5 in columns E to N
        for row in sheet.iter_rows(min_row=2, min_col=5, max_col=17):
            for cell in row:
                if isinstance(cell.value, int) and cell.value in colors:
                    fill = PatternFill(start_color=colors[cell.value], end_color=colors[cell.value], fill_type="solid")
                    cell.fill = fill
                elif cell.value == 'null':
                    fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type="solid")
                    cell.fill = fill
                elif cell.value == '1':
                    fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
                    cell.fill = fill
                elif cell.value == '2':
                    fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type="solid")
                    cell.fill = fill
                elif cell.value == '3':
                    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")
                    cell.fill = fill
                elif cell.value == '4':
                    fill = PatternFill(start_color='00B050', end_color='00B050', fill_type="solid")
                    cell.fill = fill
                elif cell.value == '5':
                    fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type="solid")
                    cell.fill = fill
                else:
                    fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")
                    cell.fill = fill

    # Save the workbook
    response = Response()
    response.headers['Content-Disposition'] = 'attachment; filename=data.xlsx'
    response.mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    workbook.save(response.stream)

    return response


# @views.route('/downloadi_csv')
# def downloadi_csv():
@views.route('/report')
def report():
    if 'studentemail' in session :
        usn = session.get('studentemail')
        user = db.users.find_one({'usn':usn })



        Assessment1 =user.get('Assessment1')
        Assessment2 =user.get('Assessment2')
        Assessment3 =user.get('Assessment3')
        Assessment4 =user.get('Assessment4')
        Assessment5 =user.get('Assessment5')
        Assessment6 =user.get('Assessment6')
        Assessment7 =user.get('Assessment7')
        Assessment8 =user.get('Assessment8')
    #    this is for test1.... 
        
        if   Assessment1 is not None and Assessment2 is  None  and Assessment3 is  None  and Assessment4 is  None  and Assessment5  is  None and Assessment6  is  None and Assessment7 is  None and Assessment8  is  None:
        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            def summary(userr):
                p=[]
                for key, value in userr.items():
                        p.append(int(value))
                return p  


            list_of_first_test_results=summary(test1_skills)


            

            communication=[list_of_first_test_results[0]]
            technical=[list_of_first_test_results[1]]
            creativity=[list_of_first_test_results[2]]
            projectmmt=[list_of_first_test_results[3]]
            timemmt=[list_of_first_test_results[4]]            
            gk=[list_of_first_test_results[5]]     
            interpersonal=[list_of_first_test_results[6]]
            resultoriented=[list_of_first_test_results[7]]
            leadership=[list_of_first_test_results[8]]
            presentation=[list_of_first_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10]]
            aptitude=[list_of_first_test_results[11]]
            # communication1 = int(user['Assessment1'].get('communication'))
            # technical1 = int(user["Assessment1"].get('technical'))
            # creativity1 = int(user['Assessment1'].get('creativity'))
            # projectmm1 = int(user['Assessment1'].get('projectmmt'))
            # timemanagement1 = int(user['Assessment1'].get('timemanagement'))
            # generalknowledge1 = int(user['Assessment1'].get('generalknowledge'))
            # interpersonal1 = int(user['Assessment1'].get('interpersonal'))
            # resultoriented1 = int(user['Assessment1'].get('resultoriented'))
            # leardership1 = int(user['Assessment1'].get('leardership'))
            # presentation1 = int(user['Assessment1'].get('presentation'))

            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/15)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/25)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/25)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/25)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/25)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/25)*5))
            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)
            
            print(aptitude)
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('report.html',user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,presentation=json.dumps(presentation),resultoriented=json.dumps(resultoriented),leadership=json.dumps(leadership),ent1=json.dumps(entreprenuerial),apt=json.dumps(aptitude),interpersonal=json.dumps(interpersonal),communication=json.dumps(communication),technical=json.dumps(technical),creativity=json.dumps(creativity),projectmmt=json.dumps(projectmmt),timemmt=json.dumps(timemmt),gk=json.dumps(gk))
        # this is for test2..  
        elif  Assessment1 is not None and Assessment2 is not None and  Assessment3 is  None and Assessment4 is  None and Assessment5 is  None and Assessment6 is  None  and Assessment7 is  None and Assessment8 is  None:
            
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            def summary(userr):
                p=[]
                for key, value in userr.items():
                    p.append(int(value))
                return p  
                        
                                

                        

          
            
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
           
         
            

            communication=[list_of_first_test_results[0],list_of_second_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4]]            
            gk=[list_of_first_test_results[5],list_of_second_test_results[5]]     
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11]]

            # communication1 = int(user['Assessment1'].get('communication'))
            # technical1 = int(user["Assessment1"].get('technical'))
            # creativity1 = int(user['Assessment1'].get('creativity'))
            # projectmm1 = int(user['Assessment1'].get('projectmmt'))
            # timemanagement1 = int(user['Assessment1'].get('timemanagement'))
            # generalknowledge1 = int(user['Assessment1'].get('generalknowledge'))
            # interpersonal1 = int(user['Assessment1'].get('interpersonal'))
            # resultoriented1 = int(user['Assessment1'].get('resultoriented'))
            # leardership1 = int(user['Assessment1'].get('leardership'))
            # presentation1 = int(user['Assessment1'].get('presentation'))

            

            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/30)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/50)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/50)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/50)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/50)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/50)*5))

            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)
           
            print(aptitude)
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('report.html',user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,presentation=json.dumps(presentation),resultoriented=json.dumps(resultoriented),leadership=json.dumps(leadership),ent1=json.dumps(entreprenuerial),apt=json.dumps(aptitude),interpersonal=json.dumps(interpersonal),communication=json.dumps(communication),technical=json.dumps(technical),creativity=json.dumps(creativity),projectmmt=json.dumps(projectmmt),timemmt=json.dumps(timemmt),gk=json.dumps(gk))
    #    this is for test3 ... 
        elif    Assessment1 is not None and Assessment2 is not None and  Assessment3 is not None and Assessment4 is  None and Assessment5 is  None and Assessment6 is  None  and Assessment7 is  None and Assessment8 is  None:


        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            
            def summary(userr):
                p=[]
                for key, value in userr.items():
                        p.append(int(value))
                return p  
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            test3_skills = user["Assessment3"]
          
            
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
            list_of_third_test_results=summary(test3_skills)
           
         
            

            communication=[list_of_first_test_results[0],list_of_second_test_results[0],list_of_third_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1],list_of_third_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2],list_of_third_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3],list_of_third_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4],list_of_third_test_results[4]]            
            gk=[list_of_first_test_results[5],list_of_second_test_results[5],list_of_third_test_results[5]]     
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6],list_of_third_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7],list_of_third_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8],list_of_third_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9],list_of_third_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10],list_of_third_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11],list_of_third_test_results[11]]
            print(aptitude)

            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/45)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/75)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/75)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/75)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/60)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/75)*5))

            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)
            
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('report.html',user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,presentation=json.dumps(presentation),resultoriented=json.dumps(resultoriented),leadership=json.dumps(leadership),ent1=json.dumps(entreprenuerial),apt=json.dumps(aptitude),interpersonal=json.dumps(interpersonal),communication=json.dumps(communication),technical=json.dumps(technical),creativity=json.dumps(creativity),projectmmt=json.dumps(projectmmt),timemmt=json.dumps(timemmt),gk=json.dumps(gk))
    #    this is for assessmet4 ....

        elif    Assessment1 is not None and Assessment2 is not None and  Assessment3 is not None and Assessment4 is not  None and Assessment5 is  None and Assessment6 is  None  and Assessment7 is  None and Assessment8 is  None:


        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            test3_skills = user["Assessment3"]
            test4_skills = user["Assessment4"]
            def summary(userr):
                        p=[]
                        for key, value in userr.items():
                                p.append(int(value))
                        return p  
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
            list_of_third_test_results=summary(test3_skills)
            list_of_fourth_test_results=summary(test4_skills)
          
          

            communication=[list_of_first_test_results[0],list_of_second_test_results[0],list_of_third_test_results[0],list_of_fourth_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1],list_of_third_test_results[1],list_of_fourth_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2],list_of_third_test_results[2],list_of_fourth_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3],list_of_third_test_results[3],list_of_fourth_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4],list_of_third_test_results[4],list_of_fourth_test_results[4]]
            gk=[list_of_first_test_results[5],list_of_second_test_results[5],list_of_third_test_results[5],list_of_fourth_test_results[5]]
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6],list_of_third_test_results[6],list_of_fourth_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7],list_of_third_test_results[7],list_of_fourth_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8],list_of_third_test_results[8],list_of_fourth_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9],list_of_third_test_results[9],list_of_fourth_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10],list_of_third_test_results[10],list_of_fourth_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11],list_of_third_test_results[11],list_of_fourth_test_results[11]]
            print(aptitude)
            
            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/60)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/100)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/100)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/100)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/80)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/100)*5))
            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)

            
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('report.html',user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,presentation=json.dumps(presentation),resultoriented=json.dumps(resultoriented),leadership=json.dumps(leadership),ent1=json.dumps(entreprenuerial),apt=json.dumps(aptitude),interpersonal=json.dumps(interpersonal),communication=json.dumps(communication),technical=json.dumps(technical),creativity=json.dumps(creativity),projectmmt=json.dumps(projectmmt),timemmt=json.dumps(timemmt),gk=json.dumps(gk))
        # this is for test5...
        elif  Assessment1 is not None and Assessment2 is not None and  Assessment3 is not None and Assessment4 is not  None and Assessment5 is not  None and Assessment6 is  None  and Assessment7 is  None and Assessment8 is  None:


        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            test3_skills = user["Assessment3"]
            test4_skills = user["Assessment4"]
            test5_skills = user["Assessment5"]
            def summary(userr):
                p=[]
                for key, value in userr.items():
                        p.append(int(value))
                return p  
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
            list_of_third_test_results=summary(test3_skills)
            list_of_fourth_test_results=summary(test4_skills)
            list_of_fifth_test_results=summary(test5_skills)
          
          

            communication=[list_of_first_test_results[0],list_of_second_test_results[0],list_of_third_test_results[0],list_of_fourth_test_results[0],list_of_fifth_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1],list_of_third_test_results[1],list_of_fourth_test_results[1],list_of_fifth_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2],list_of_third_test_results[2],list_of_fourth_test_results[2],list_of_fifth_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3],list_of_third_test_results[3],list_of_fourth_test_results[3],list_of_fifth_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4],list_of_third_test_results[4],list_of_fourth_test_results[4],list_of_fifth_test_results[4]]
            gk=[list_of_first_test_results[5],list_of_second_test_results[5],list_of_third_test_results[5],list_of_fourth_test_results[5],list_of_fifth_test_results[5]]
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6],list_of_third_test_results[6],list_of_fourth_test_results[6],list_of_fifth_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7],list_of_third_test_results[7],list_of_fourth_test_results[7],list_of_fifth_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8],list_of_third_test_results[8],list_of_fourth_test_results[8],list_of_fifth_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9],list_of_third_test_results[9],list_of_fourth_test_results[9],list_of_fifth_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10],list_of_third_test_results[10],list_of_fourth_test_results[10],list_of_fifth_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11],list_of_third_test_results[11],list_of_fourth_test_results[11],list_of_fifth_test_results[11]]
            print(aptitude)
            
            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/75)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/125)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/125)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/125)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/100)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/125)*5))
    
            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)

            
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('report.html',user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,presentation=json.dumps(presentation),resultoriented=json.dumps(resultoriented),leadership=json.dumps(leadership),ent1=json.dumps(entreprenuerial),apt=json.dumps(aptitude),interpersonal=json.dumps(interpersonal),communication=json.dumps(communication),technical=json.dumps(technical),creativity=json.dumps(creativity),projectmmt=json.dumps(projectmmt),timemmt=json.dumps(timemmt),gk=json.dumps(gk))
    #    this is for test6... 
        elif  Assessment1 is not None and Assessment2 is not None and  Assessment3 is not None and Assessment4 is not  None and Assessment5 is not  None and Assessment6 is  not None  and Assessment7 is  None and Assessment8 is  None:


        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            test3_skills = user["Assessment3"]
            test4_skills = user["Assessment4"]
            test5_skills = user["Assessment5"]
            test6_skills = user["Assessment6"]
            def summary(userr):
                p=[]
                for key, value in userr.items():
                        p.append(int(value))
                return p  
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
            list_of_third_test_results=summary(test3_skills)
            list_of_fourth_test_results=summary(test4_skills)
            list_of_fifth_test_results=summary(test5_skills)
            list_of_sixth_test_results=summary(test6_skills)
            # communication1 = int(user['Assessment1'].get('communication'))
            # technical1 = int(user["Assessment1"].get('technical'))
            # creativity1 = int(user['Assessment1'].get('creativity'))
            # projectmm1 = int(user['Assessment1'].get('projectmmt'))
            # timemanagement1 = int(user['Assessment1'].get('timemanagement'))
            # generalknowledge1 = int(user['Assessment1'].get('generalknowledge'))
            # interpersonal1 = int(user['Assessment1'].get('interpersonal'))
            # resultoriented1 = int(user['Assessment1'].get('resultoriented'))
            # leardership1 = int(user['Assessment1'].get('leardership'))
            # presentation1 = int(user['Assessment1'].get('presentation'))
            communication=[list_of_first_test_results[0],list_of_second_test_results[0],list_of_third_test_results[0],list_of_fourth_test_results[0],list_of_fifth_test_results[0],list_of_sixth_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1],list_of_third_test_results[1],list_of_fourth_test_results[1],list_of_fifth_test_results[1],list_of_sixth_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2],list_of_third_test_results[2],list_of_fourth_test_results[2],list_of_fifth_test_results[2],list_of_sixth_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3],list_of_third_test_results[3],list_of_fourth_test_results[3],list_of_fifth_test_results[3],list_of_sixth_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4],list_of_third_test_results[4],list_of_fourth_test_results[4],list_of_fifth_test_results[4],list_of_sixth_test_results[4]]
            gk=[list_of_first_test_results[5],list_of_second_test_results[5],list_of_third_test_results[5],list_of_fourth_test_results[5],list_of_fifth_test_results[5],list_of_sixth_test_results[5]]
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6],list_of_third_test_results[6],list_of_fourth_test_results[6],list_of_fifth_test_results[6],list_of_sixth_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7],list_of_third_test_results[7],list_of_fourth_test_results[7],list_of_fifth_test_results[7],list_of_sixth_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8],list_of_third_test_results[8],list_of_fourth_test_results[8],list_of_fifth_test_results[8],list_of_sixth_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9],list_of_third_test_results[9],list_of_fourth_test_results[9],list_of_fifth_test_results[9],list_of_sixth_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10],list_of_third_test_results[10],list_of_fourth_test_results[10],list_of_fifth_test_results[10],list_of_sixth_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11],list_of_third_test_results[11],list_of_fourth_test_results[11],list_of_fifth_test_results[11],list_of_sixth_test_results[11]]
            print(sum(technical))
            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/90)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/150)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/150)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/150)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/120)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/150)*5))
            print(customeri)
            print(projecti)
            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)

            
            

            
            print('++++')
            return render_template('report.html',user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,presentation=json.dumps(presentation),resultoriented=json.dumps(resultoriented),leadership=json.dumps(leadership),ent1=json.dumps(entreprenuerial),apt=json.dumps(aptitude),interpersonal=json.dumps(interpersonal),communication=json.dumps(communication),technical=json.dumps(technical),creativity=json.dumps(creativity),projectmmt=json.dumps(projectmmt),timemmt=json.dumps(timemmt),gk=json.dumps(gk))
        # this is for test7
        elif  Assessment1 is not None and Assessment2 is not None and  Assessment3 is not None and Assessment4 is not  None and Assessment5 is not None and Assessment6 is not None  and Assessment7 is not None and Assessment8 is  None:


        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            test3_skills = user["Assessment3"]
            test4_skills = user["Assessment4"]
            test5_skills = user["Assessment5"]
            test6_skills = user["Assessment6"]
            test7_skills = user["Assessment7"]
            def summary(userr):
                p=[]
                for key, value in userr.items():
                        p.append(int(value))
                return p  
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
            list_of_third_test_results=summary(test3_skills)
            list_of_fourth_test_results=summary(test4_skills)
            list_of_fifth_test_results=summary(test5_skills)
            list_of_sixth_test_results=summary(test6_skills)
            list_of_seventh_test_results=summary(test7_skills)
           

            communication=[list_of_first_test_results[0],list_of_second_test_results[0],list_of_third_test_results[0],list_of_fourth_test_results[0],list_of_fifth_test_results[0],list_of_sixth_test_results[0],list_of_seventh_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1],list_of_third_test_results[1],list_of_fourth_test_results[1],list_of_fifth_test_results[1],list_of_sixth_test_results[1],list_of_seventh_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2],list_of_third_test_results[2],list_of_fourth_test_results[2],list_of_fifth_test_results[2],list_of_sixth_test_results[2],list_of_seventh_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3],list_of_third_test_results[3],list_of_fourth_test_results[3],list_of_fifth_test_results[3],list_of_sixth_test_results[3],list_of_seventh_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4],list_of_third_test_results[4],list_of_fourth_test_results[4],list_of_fifth_test_results[4],list_of_sixth_test_results[4],list_of_seventh_test_results[4]]
            gk=[list_of_first_test_results[5],list_of_second_test_results[5],list_of_third_test_results[5],list_of_fourth_test_results[5],list_of_fifth_test_results[5],list_of_sixth_test_results[5],list_of_seventh_test_results[5]]
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6],list_of_third_test_results[6],list_of_fourth_test_results[6],list_of_fifth_test_results[6],list_of_sixth_test_results[6],list_of_seventh_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7],list_of_third_test_results[7],list_of_fourth_test_results[7],list_of_fifth_test_results[7],list_of_sixth_test_results[7],list_of_seventh_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8],list_of_third_test_results[8],list_of_fourth_test_results[8],list_of_fifth_test_results[8],list_of_sixth_test_results[8],list_of_seventh_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9],list_of_third_test_results[9],list_of_fourth_test_results[9],list_of_fifth_test_results[9],list_of_sixth_test_results[9],list_of_seventh_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10],list_of_third_test_results[10],list_of_fourth_test_results[10],list_of_fifth_test_results[10],list_of_sixth_test_results[10],list_of_seventh_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11],list_of_third_test_results[11],list_of_fourth_test_results[11],list_of_fifth_test_results[11],list_of_sixth_test_results[11],list_of_seventh_test_results[11]]
            print(aptitude)
            
            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/105)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/175)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/175)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/175)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/140)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/175)*5))
            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)

            
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('report.html',user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,presentation=json.dumps(presentation),resultoriented=json.dumps(resultoriented),leadership=json.dumps(leadership),ent1=json.dumps(entreprenuerial),apt=json.dumps(aptitude),interpersonal=json.dumps(interpersonal),communication=json.dumps(communication),technical=json.dumps(technical),creativity=json.dumps(creativity),projectmmt=json.dumps(projectmmt),timemmt=json.dumps(timemmt),gk=json.dumps(gk))
        # this is for test8
        elif Assessment8 is not None:
        
            usn = session.get('studentemail')
            user = db.users.find_one({'usn':usn })
            test1_skills = user["Assessment1"]
            test2_skills = user["Assessment2"]
            test3_skills = user["Assessment3"]
            test4_skills = user["Assessment4"]
            test5_skills = user["Assessment5"]
            test6_skills = user["Assessment6"]
            test7_skills = user["Assessment7"]
            test8_skills = user["Assessment8"]
            def summary(userr):
                p=[]
                for key, value in userr.items():
                        p.append(int(value))
                return p  
            list_of_second_test_results=summary(test2_skills)            
            list_of_first_test_results=summary(test1_skills)
            list_of_third_test_results=summary(test3_skills)
            list_of_fourth_test_results=summary(test4_skills)
            list_of_fifth_test_results=summary(test5_skills)
            list_of_sixth_test_results=summary(test6_skills)
            list_of_seventh_test_results=summary(test7_skills)
            print(list_of_sixth_test_results)
            list_of_eight_test_results=summary(test8_skills)
            print(list_of_eight_test_results)
            list_of_assessments=["Assessment1","Assessment2","Assessment3","Assessment4","Assessment5","Assessment6","Assessment7","Assessment8"]
            print(list_of_assessments)

            communication=[list_of_first_test_results[0],list_of_second_test_results[0],list_of_third_test_results[0],list_of_fourth_test_results[0],list_of_fifth_test_results[0],list_of_sixth_test_results[0],list_of_seventh_test_results[0],list_of_eight_test_results[0]]
            technical=[list_of_first_test_results[1],list_of_second_test_results[1],list_of_third_test_results[1],list_of_fourth_test_results[1],list_of_fifth_test_results[1],list_of_sixth_test_results[1],list_of_seventh_test_results[1],list_of_eight_test_results[1]]
            creativity=[list_of_first_test_results[2],list_of_second_test_results[2],list_of_third_test_results[2],list_of_fourth_test_results[2],list_of_fifth_test_results[2],list_of_sixth_test_results[2],list_of_seventh_test_results[2],list_of_eight_test_results[2]]
            projectmmt=[list_of_first_test_results[3],list_of_second_test_results[3],list_of_third_test_results[3],list_of_fourth_test_results[3],list_of_fifth_test_results[3],list_of_sixth_test_results[3],list_of_seventh_test_results[3],list_of_eight_test_results[3]]
            timemmt=[list_of_first_test_results[4],list_of_second_test_results[4],list_of_third_test_results[4],list_of_fourth_test_results[4],list_of_fifth_test_results[4],list_of_sixth_test_results[4],list_of_seventh_test_results[4],list_of_eight_test_results[4]]
            gk=[list_of_first_test_results[5],list_of_second_test_results[5],list_of_third_test_results[5],list_of_fourth_test_results[5],list_of_fifth_test_results[5],list_of_sixth_test_results[5],list_of_seventh_test_results[5],list_of_eight_test_results[5]]
            interpersonal=[list_of_first_test_results[6],list_of_second_test_results[6],list_of_third_test_results[6],list_of_fourth_test_results[6],list_of_fifth_test_results[6],list_of_sixth_test_results[6],list_of_seventh_test_results[6],list_of_eight_test_results[6]]
            resultoriented=[list_of_first_test_results[7],list_of_second_test_results[7],list_of_third_test_results[7],list_of_fourth_test_results[7],list_of_fifth_test_results[7],list_of_sixth_test_results[7],list_of_seventh_test_results[7],list_of_eight_test_results[7]]
            leadership=[list_of_first_test_results[8],list_of_second_test_results[8],list_of_third_test_results[8],list_of_fourth_test_results[8],list_of_fifth_test_results[8],list_of_sixth_test_results[8],list_of_seventh_test_results[8],list_of_eight_test_results[8]]
            presentation=[list_of_first_test_results[9],list_of_second_test_results[9],list_of_third_test_results[9],list_of_fourth_test_results[9],list_of_fifth_test_results[9],list_of_sixth_test_results[9],list_of_seventh_test_results[9],list_of_eight_test_results[9]]
            entreprenuerial=[list_of_first_test_results[10],list_of_second_test_results[10],list_of_third_test_results[10],list_of_fourth_test_results[10],list_of_fifth_test_results[10],list_of_sixth_test_results[10],list_of_seventh_test_results[10],list_of_eight_test_results[10]]
            aptitude=[list_of_first_test_results[11],list_of_second_test_results[11],list_of_third_test_results[11],list_of_fourth_test_results[11],list_of_fifth_test_results[11],list_of_sixth_test_results[11],list_of_seventh_test_results[11],list_of_eight_test_results[11]]
            print(aptitude)
            
            strongi = int((((sum(technical)+sum(creativity)+sum(resultoriented)))/120)*5)
            leaderi = int((((sum(projectmmt)+sum(leadership)+sum(timemmt)+sum(resultoriented)+sum(communication)))/200)*5)
            customeri = int((((sum(presentation)+sum(gk)+sum(timemmt)+sum(interpersonal)+sum(communication)))/200)*5)
            projecti = int((((sum(projectmmt)+sum(resultoriented)+sum(timemmt)+sum(technical)+sum(communication))/200)*5))
            designi = int((((sum(leadership)+sum(resultoriented)+sum(creativity)+sum(technical))/160)*5))
            marketi = int(((( sum(presentation)+sum(resultoriented)+sum(communication)+sum(interpersonal)+sum(gk))/200)*5))
            def st(strongi):
                if strongi >= 4:
                    return "Strong in Tech"
                else:
                    return "Generic"
            strong = st(strongi)
            def lr(leaderi):
                if leaderi >= 4:
                    return "Leadership Roles"
                else:
                    return "Generic"
            leader = lr(leaderi)
            def cfr(customei):
                if  customeri >3 :
                    return "Customer Facing Roles"
                else:
                    return "Generic"
            customer = cfr(customeri)
            def pm(customer):
                if  projecti >3 :
                    return "Project Management "
                else:
                    return "None"
            project = pm(projecti)
            def df(designi):
                if designi >=3  :
                    return "Design Profile "
                else:
                    return "None"
            design = df(designi)
            def mr(marketi):
                if marketi >= 3  :
                    return "Marketing Role "
                else:
                    return "None"
            market = mr(projecti)

            
            print(entreprenuerial)
            print(gk)
            print('++++')
            return render_template('report.html',user=user,design=design,market=market,strong=strong,leader=leader,customer=customer,project=project,presentation=json.dumps(presentation),resultoriented=json.dumps(resultoriented),leadership=json.dumps(leadership),ent1=json.dumps(entreprenuerial),apt=json.dumps(aptitude),interpersonal=json.dumps(interpersonal),communication=json.dumps(communication),technical=json.dumps(technical),creativity=json.dumps(creativity),projectmmt=json.dumps(projectmmt),timemmt=json.dumps(timemmt),gk=json.dumps(gk))


@views.route('/resume')
def resume():
    if 'studentemail' in session :
        usn = session.get('studentemail')
        user = db.users.find_one({'usn':usn })
        username = user["personal"].get('username')
        email = user["personal"].get('email')
        college = user["personal"].get('college')
        phone = user["personal"].get('phone')
        numberofcertifications = user["personal"].get('numberofcertifications')
        technicalskills = user["personal"].get('technicalskills')
        softskills = user["personal"].get('softskills')
        tenthboard = user["personal"].get('tenthboard')
        tenthmarks = user["personal"].get('tenthmarks')
        twelvethboard = user["personal"].get('twelvethboard')
        twelvethmarks = user["personal"].get('twelvethmarks')
        ugaverage = user["personal"].get('ugaverage')
        linkdin = user["personal"].get('linkdin')
        github = user["personal"].get('github')
        skill1 = user["personal"].get('skill1')
        skill2 = user["personal"].get('skill2')
        skill3 = user["personal"].get('skill3')
        skill4 = user["personal"].get('skill4')
        skill5 = user["personal"].get('skill5')
        certification1 = user["personal"].get('certification1')
        certification2 = user["personal"].get('certification2')
        certification3= user["personal"].get('certification3')
        certification4 = user["personal"].get('certification4')
        certification5 = user["personal"].get('certification5')
        
        # for key,value in p.items():
        #     if value=="dsce":
        #         print("heyy well done ") 
        #     else:
        #         print('better luck next time')
        

        
        title="dashboard"

    return render_template('resume.html',user=user)


@views.route('/score',methods=['GET','POST'])
def score():
    usn = session.get('studentname')
    user = db.users.find_one({'usn':usn })
    username = user["personal"].get('username')
    userusn = user["personal"].get('usn')

    title="marks enter"
    if  'teacheremail' in session:

        if request.method == 'POST':

                test= request.form.get('test')
                form1= request.form
                c1=form1.get('communication')
                print(c1)
                # t1=form1.get('')
                communication= request.form.get('communication')
                technical= request.form.get('technical')
                creativity= request.form.get('creativity')
                projectmmt= request.form.get('projectmmt')
                timemanagement= request.form.get('timemanagement')
                generalknowledge= request.form.get('generalknowledge')
                interpersonal= request.form.get('interpersonal')
                resultoriented= request.form.get('resultoriented')
                leardership= request.form.get('leardership')
                presentation= request.form.get('presentation')
                aptitude= request.form.get('aptitude')
                entrepreneur= request.form.get('entrepreneur')
                
                print(leardership)
                print(presentation)
                print(communication)

                # user = db.users.find_one({'username':username })
                # email = user['email']
                # db.marks.insert_one({"communication": communication})
                # if test == "Test1":
                try:
                    db.users.update_one(
                        {"usn": usn},
                        {
                            "$set": {
                                test: {
                                    "communication": communication,
                                    "technical": technical,
                                    "creativity": creativity,
                                    "projectmmt": projectmmt,
                                    "timemanagement": timemanagement,
                                    "generalknowledge": generalknowledge,
                                    "interpersonal": interpersonal,
                                    "resultoriented": resultoriented,
                                    "leardership": leardership,
                                    "presentation": presentation,
                                    "aptitude": aptitude,
                                    "entrepreneur": entrepreneur
                                }
                            }
                        }
                    )
                    flash('Marks entered successfully', category='success')
                    return redirect(url_for('views.dashboard'))

                        
                except PyMongoError as e:
                    flash(f'Error: {str(e)}', category='error')
                    # Handle the error accordingly, such as logging it or displaying an error message to the user
                    return redirect(url_for('views.score'))
            
                    
        
        # print(studentname)
        # return render_template("marksenter.html",title=title,studentname=studentname)
        return render_template("score.html",username=username,usn=usn)
    return redirect(url_for('auth.login'))

























